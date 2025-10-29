from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_caching import Cache
from flask_oauthlib.client import OAuth
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect
from flask_jwt_extended import JWTManager, create_access_token, create_refresh_token, jwt_required, get_jwt_identity, get_jwt
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import secrets
import bleach
import pyotp
import qrcode
from io import BytesIO
import base64
import csv
import hashlib
import uuid
from functools import wraps
import re
from dotenv import load_dotenv
from config import config
from api_routes import api

# Carregar variáveis de ambiente
load_dotenv()

# Funções de validação
def validate_cpf(cpf):
    """Valida CPF brasileiro"""
    if not cpf:
        return False
    
    # Remove caracteres não numéricos
    cpf = re.sub(r'[^0-9]', '', cpf)
    
    # Verifica se tem 11 dígitos
    if len(cpf) != 11:
        return False
    
    # Verifica se todos os dígitos são iguais
    if cpf == cpf[0] * 11:
        return False
    
    # Calcula o primeiro dígito verificador
    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    resto = soma % 11
    digito1 = 0 if resto < 2 else 11 - resto
    
    # Calcula o segundo dígito verificador
    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    resto = soma % 11
    digito2 = 0 if resto < 2 else 11 - resto
    
    return cpf[-2:] == f"{digito1}{digito2}"

def validate_cnpj(cnpj):
    """Valida CNPJ brasileiro"""
    if not cnpj:
        return False
    
    # Remove caracteres não numéricos
    cnpj = re.sub(r'[^0-9]', '', cnpj)
    
    # Verifica se tem 14 dígitos
    if len(cnpj) != 14:
        return False
    
    # Verifica se todos os dígitos são iguais
    if cnpj == cnpj[0] * 14:
        return False
    
    # Calcula o primeiro dígito verificador
    multiplicadores1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    soma = sum(int(cnpj[i]) * multiplicadores1[i] for i in range(12))
    resto = soma % 11
    digito1 = 0 if resto < 2 else 11 - resto
    
    # Calcula o segundo dígito verificador
    multiplicadores2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    soma = sum(int(cnpj[i]) * multiplicadores2[i] for i in range(13))
    resto = soma % 11
    digito2 = 0 if resto < 2 else 11 - resto
    
    return cnpj[-2:] == f"{digito1}{digito2}"

def validate_email(email):
    """Valida formato de email"""
    if not email:
        return True  # Email é opcional
    
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def format_phone(phone):
    """Formata telefone brasileiro"""
    if not phone:
        return phone
    
    # Remove caracteres não numéricos
    phone = re.sub(r'[^0-9]', '', phone)
    
    # Formata baseado no tamanho
    if len(phone) == 11:  # Celular com DDD
        return f"({phone[:2]}) {phone[2:7]}-{phone[7:]}"
    elif len(phone) == 10:  # Telefone fixo com DDD
        return f"({phone[:2]}) {phone[2:6]}-{phone[6:]}"
    elif len(phone) == 9:  # Celular sem DDD
        return f"{phone[:5]}-{phone[5:]}"
    elif len(phone) == 8:  # Telefone fixo sem DDD
        return f"{phone[:4]}-{phone[4:]}"
    
    return phone

def format_cpf_cnpj(doc):
    """Formata CPF ou CNPJ"""
    if not doc:
        return doc
    
    # Remove caracteres não numéricos
    doc = re.sub(r'[^0-9]', '', doc)
    
    if len(doc) == 11:  # CPF
        return f"{doc[:3]}.{doc[3:6]}.{doc[6:9]}-{doc[9:]}"
    elif len(doc) == 14:  # CNPJ
        return f"{doc[:2]}.{doc[2:5]}.{doc[5:8]}/{doc[8:12]}-{doc[12:]}"
    
    return doc

def get_cep_data(cep):
    """Busca dados do CEP via API"""
    try:
        # Remove caracteres não numéricos
        cep = re.sub(r'[^0-9]', '', cep)
        
        if len(cep) != 8:
            return None
        
        # Usa a API ViaCEP
        response = requests.get(f'https://viacep.com.br/ws/{cep}/json/', timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            if 'erro' not in data:
                return {
                    'logradouro': data.get('logradouro', ''),
                    'bairro': data.get('bairro', ''),
                    'cidade': data.get('localidade', ''),
                    'uf': data.get('uf', ''),
                    'cep': data.get('cep', '')
                }
    except Exception as e:
        print(f"Erro ao buscar CEP: {e}")
    
    return None
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# Configurações
config_name = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config[config_name])

# Configurações do Google OAuth
app.config['GOOGLE_CLIENT_ID'] = os.environ.get('GOOGLE_CLIENT_ID', '')
app.config['GOOGLE_CLIENT_SECRET'] = os.environ.get('GOOGLE_CLIENT_SECRET', '')

# Configurações para produção (Render)
if os.environ.get('RENDER'):
    app.config['PREFERRED_URL_SCHEME'] = 'https'

db = SQLAlchemy(app)

# Filtro Jinja2 para formatar datas

def datetime_filter(value, fmt='short'):
    try:
        from datetime import datetime, date
        if not value:
            return ''
        if isinstance(value, str):
            try:
                value_parsed = datetime.fromisoformat(value)
                value = value_parsed
            except Exception:
                return value
        if isinstance(value, date) and not isinstance(value, datetime):
            if fmt in ('short', 'date'):
                return value.strftime('%d/%m/%Y')
            elif fmt in ('long', 'full'):
                return value.strftime('%d/%m/%Y')
            elif fmt == 'iso':
                return value.isoformat()
            else:
                return value.strftime('%d/%m/%Y')
        if isinstance(value, datetime):
            if fmt == 'short':
                return value.strftime('%d/%m/%Y %H:%M')
            elif fmt in ('long', 'full'):
                return value.strftime('%d/%m/%Y %H:%M:%S')
            elif fmt == 'date':
                return value.strftime('%d/%m/%Y')
            elif fmt == 'time':
                return value.strftime('%H:%M')
            elif fmt == 'iso':
                return value.isoformat(sep=' ')
            else:
                return value.strftime('%d/%m/%Y %H:%M')
        return str(value)
    except Exception:
        return str(value)

app.jinja_env.filters['datetime'] = datetime_filter
migrate = Migrate(app, db)

# Configuração do Cache
try:
    cache = Cache(app, config={
        'CACHE_TYPE': 'simple',  # Para desenvolvimento, use 'redis' em produção
        'CACHE_DEFAULT_TIMEOUT': 300  # 5 minutos
    })
except Exception as e:
    print(f"AVISO: Erro ao configurar cache: {e}")
    cache = None

# ========== CONFIGURAÇÕES DE SEGURANÇA APRIMORADAS ==========

# Importar middleware de segurança
try:
    from security_middleware import SecurityMiddleware, require_https, require_csrf, rate_limit
    from security_validators import SecurityValidators
    print("Middleware de seguranca carregado")
except ImportError as e:
    print(f"AVISO: Middleware de seguranca nao disponivel: {e}")
    SecurityMiddleware = None
    SecurityValidators = None

# Configuração de Segurança
try:
    csrf = CSRFProtect(app)
    print("CSRF Protection habilitado")
except Exception as e:
    print(f"AVISO: Erro ao configurar CSRF: {e}")
    csrf = None

# Inicializar middleware de segurança
if SecurityMiddleware:
    security_middleware = SecurityMiddleware(app)
    print("Middleware de seguranca inicializado")
else:
    security_middleware = None

# CSRF protection aplicado apenas às rotas web
# Rotas da API usam @csrf.exempt individualmente

# Configuração adicional do CSRF
app.config['WTF_CSRF_TIME_LIMIT'] = 3600  # 1 hora
app.config['WTF_CSRF_SSL_STRICT'] = os.environ.get('WTF_CSRF_SSL_STRICT', 'false').lower() == 'true'

# Headers de segurança
@app.after_request
def add_security_headers(response):
    """Adiciona headers de segurança"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    
    # HSTS apenas em produção
    if os.environ.get('RENDER') or os.environ.get('FLASK_ENV') == 'production':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    return response

# Função para aplicar CSRF exempt condicionalmente
def csrf_exempt_if_needed(func):
    """Aplica @csrf.exempt apenas se csrf estiver disponível"""
    if csrf:
        return csrf.exempt(func)
    return func

# Configuração do Rate Limiting (mais tolerante em desenvolvimento)
try:
    dev_mode = os.environ.get('FLASK_ENV', 'development').lower() == 'development' or os.environ.get('FLASK_DEBUG') == '1'
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        # Em desenvolvimento, desabilita limites globais; em produção, aumenta tolerância e habilita cabeçalhos
        default_limits=[] if dev_mode else ["1000 per day", "200 per hour", "40 per minute"],
        storage_uri="memory://",
        headers_enabled=True
    )
except Exception as e:
    print(f"AVISO: Erro ao configurar Rate Limiting: {e}")
    limiter = None

# Configuração JWT para Mobile
try:
    app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', secrets.token_hex(32))
    app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=1)
    app.config['JWT_REFRESH_TOKEN_EXPIRES'] = timedelta(days=30)
    jwt = JWTManager(app)
except Exception as e:
    print(f"AVISO: Erro ao configurar JWT: {e}")
    jwt = None

# Configuração CORS para Mobile
try:
    CORS(app, origins=['*'], supports_credentials=True)
except Exception as e:
    print(f"AVISO: Erro ao configurar CORS: {e}")

# Registrar blueprint da API
app.register_blueprint(api)
print("API routes registradas com sucesso!")

# Configuração de sessão mais segura
app.config.update(
    SESSION_COOKIE_SECURE=True if os.environ.get('RENDER') else False,  # HTTPS em produção
    SESSION_COOKIE_HTTPONLY=True,  # Previne acesso via JavaScript
    SESSION_COOKIE_SAMESITE='Lax',  # Proteção CSRF
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),  # Sessão expira em 8 horas
    SECRET_KEY=os.environ.get('SECRET_KEY', secrets.token_hex(32))  # Chave secreta segura
)

# Configuração do OAuth (apenas se as credenciais estiverem configuradas)
try:
    oauth = OAuth(app)
    google = None

    if app.config['GOOGLE_CLIENT_ID'] and app.config['GOOGLE_CLIENT_SECRET']:
        google = oauth.remote_app(
            'google',
            consumer_key=app.config['GOOGLE_CLIENT_ID'],
            consumer_secret=app.config['GOOGLE_CLIENT_SECRET'],
            request_token_params={'scope': 'email profile'},
            base_url='https://www.googleapis.com/oauth2/v1/',
            request_token_url=None,
            access_token_method='POST',
            access_token_url='https://accounts.google.com/o/oauth2/token',
            authorize_url='https://accounts.google.com/o/oauth2/auth',
        )
    else:
        print("AVISO: Google OAuth nao configurado. Configure GOOGLE_CLIENT_ID e GOOGLE_CLIENT_SECRET no arquivo .env")
except Exception as e:
    print(f"AVISO: Erro ao configurar OAuth: {e}")
    oauth = None
    google = None

# Decorator para verificar login
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        # Enforce sessão única
        try:
            user = User.query.get(session.get('user_id'))
            sess_token = session.get('active_session_id')
            if user and user.active_session_id and sess_token != user.active_session_id:
                flash('Sua sessão foi substituída por um novo login. Faça login novamente.', 'warning')
                session.clear()
                return redirect(url_for('login'))
        except Exception:
            pass
        return f(*args, **kwargs)
    return decorated_function

# Decorator seguro para rate limiting
def safe_rate_limit(limit):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Em desenvolvimento, não aplica rate limit para evitar bloqueios enquanto testa
            if limiter and not (os.environ.get('FLASK_ENV', 'development').lower() == 'development' or os.environ.get('FLASK_DEBUG') == '1'):
                return limiter.limit(limit)(f)(*args, **kwargs)
            else:
                return f(*args, **kwargs)
        return decorated_function
    return decorator

# Tratamento amigável para erros de rate limit (HTTP 429)
@app.errorhandler(429)
def ratelimit_handler(e):
    retry = 60
    try:
        retry = int(getattr(e, 'description', '').split(' ')[-1])
    except Exception:
        pass
    msg = 'Muitas requisições. Aguarde alguns segundos e tente novamente.'
    if request.accept_mimetypes.best == 'application/json':
        return jsonify({
            'success': False,
            'error': 'too_many_requests',
            'message': msg,
            'retry_after': retry
        }), 429
    flash(msg, 'warning')
    return redirect(request.referrer or url_for('index'))

# Função para obter configurações do usuário
def get_user_settings(user_id):
    settings = UserSettings.query.filter_by(user_id=user_id).first()
    if not settings:
        settings = UserSettings(user_id=user_id)
        db.session.add(settings)
        db.session.commit()
    return settings

# Função para atualizar configurações do usuário
def update_user_settings(user_id, **kwargs):
    settings = get_user_settings(user_id)
    for key, value in kwargs.items():
        if hasattr(settings, key):
            setattr(settings, key, value)
    settings.updated_at = datetime.utcnow()
    db.session.commit()
    return settings

# Helper de autorização
def is_admin():
    return session.get('role') == 'admin'

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not is_admin():
            flash('Acesso restrito ao administrador.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# ===== Dispositivo confiável e sessão única =====
def get_device_identifier(req: request) -> str:
    header_id = req.headers.get('X-Device-ID') or req.headers.get('X-Device-Id')
    if header_id:
        return header_id.strip()
    cookie_id = req.cookies.get('device_id')
    if cookie_id:
        return cookie_id.strip()
    user_agent = req.headers.get('User-Agent', '')
    accept_lang = req.headers.get('Accept-Language', '')
    remote_ip = req.headers.get('X-Forwarded-For', req.remote_addr or '')
    raw = f"{user_agent}|{accept_lang}|{remote_ip}|{os.environ.get('DEVICE_SALT','')}"
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()[:64]

def set_device_cookie(resp, device_id: str):
    try:
        resp.set_cookie('device_id', device_id, max_age=60*60*24*365, httponly=True, samesite='Lax')
    except Exception:
        pass

def issue_session_token(user) -> str:
    token = uuid.uuid4().hex
    user.active_session_id = token
    user.active_session_updated_at = datetime.utcnow()
    return token

def serialize_model(instance):
    data = {}
    for column in instance.__table__.columns:
        value = getattr(instance, column.name)
        if isinstance(value, datetime):
            value = value.isoformat()
        data[column.name] = value
    return data

# ========== FUNÇÕES DE SEGURANÇA ==========

def sanitize_input(text):
    """Sanitiza entrada do usuário para prevenir XSS usando validadores de segurança"""
    if SecurityValidators:
        return SecurityValidators.sanitize_input(text)
    else:
        # Fallback para sanitização básica
        if not text:
            return text
        
        # Remove tags HTML maliciosas
        allowed_tags = ['b', 'i', 'u', 'strong', 'em', 'br']
        allowed_attributes = {}
        
        return bleach.clean(text, tags=allowed_tags, attributes=allowed_attributes, strip=True)

def validate_email(email):
    """Valida formato de email"""
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_password_strength(password):
    """Valida força da senha usando validadores de segurança"""
    if SecurityValidators:
        return SecurityValidators.validate_password_strength(password)
    else:
        # Fallback para validação básica
        if len(password) < 8:
            return False, "Senha deve ter pelo menos 8 caracteres"
        
        if not any(c.isupper() for c in password):
            return False, "Senha deve conter pelo menos uma letra maiúscula"
        
        if not any(c.islower() for c in password):
            return False, "Senha deve conter pelo menos uma letra minúscula"
        
        if not any(c.isdigit() for c in password):
            return False, "Senha deve conter pelo menos um número"
        
        # Validar caracteres especiais
        special_chars = "!@#$%^&*()_+-=[]{}|;':\",./<>?"
        if not any(c in special_chars for c in password):
            return False, "Senha deve conter pelo menos um caractere especial (!@#$%^&*)"
        
        return True, "Senha válida"

def generate_2fa_secret():
    """Gera chave secreta para 2FA"""
    return pyotp.random_base32()

def generate_2fa_qr_code(user_email, secret):
    """Gera QR code para 2FA"""
    totp_uri = pyotp.totp.TOTP(secret).provisioning_uri(
        name=user_email,
        issuer_name="SaaS Sistema"
    )
    
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(totp_uri)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    return base64.b64encode(buffer.getvalue()).decode()

def verify_2fa_token(secret, token):
    """Verifica token 2FA"""
    totp = pyotp.TOTP(secret)
    return totp.verify(token, valid_window=1)

def generate_backup_codes():
    """Gera códigos de backup para 2FA"""
    import json
    codes = [secrets.token_hex(4).upper() for _ in range(10)]
    return json.dumps(codes)

def is_account_locked(user):
    """Verifica se conta está bloqueada"""
    if user.locked_until and user.locked_until > datetime.utcnow():
        return True
    return False

def increment_failed_login(user):
    """Incrementa tentativas de login falhadas"""
    user.failed_login_attempts += 1
    
    # Bloquear conta após 5 tentativas falhadas por 30 minutos
    if user.failed_login_attempts >= 5:
        user.locked_until = datetime.utcnow() + timedelta(minutes=30)
    
    db.session.commit()

def reset_failed_login(user):
    """Reseta tentativas de login falhadas"""
    user.failed_login_attempts = 0
    user.locked_until = None
    user.last_login = datetime.utcnow()
    db.session.commit()

# Modelos do Banco de Dados
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(120), nullable=True)  # Pode ser None para usuários Google
    empresa = db.Column(db.String(100), nullable=False, index=True)
    role = db.Column(db.String(20), default='user', nullable=False, index=True)  # 'user' | 'admin'
    google_id = db.Column(db.String(100), unique=True, nullable=True, index=True)  # ID do Google
    avatar_url = db.Column(db.String(200), nullable=True)  # URL do avatar do Google
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    # Dispositivo confiável e sessão única
    device_fingerprint = db.Column(db.String(128), nullable=True, index=True)
    device_last_updated = db.Column(db.DateTime, nullable=True)
    active_session_id = db.Column(db.String(64), nullable=True, index=True)
    active_session_updated_at = db.Column(db.DateTime, nullable=True)
    
    # Campos para 2FA
    two_factor_enabled = db.Column(db.Boolean, default=False)
    two_factor_secret = db.Column(db.String(32), nullable=True)
    backup_codes = db.Column(db.Text, nullable=True)  # JSON string dos códigos de backup
    last_login = db.Column(db.DateTime, nullable=True)
    failed_login_attempts = db.Column(db.Integer, default=0)
    locked_until = db.Column(db.DateTime, nullable=True)
    
    # Relacionamentos
    produtos = db.relationship('Produto', backref='usuario', lazy=True)
    vendas = db.relationship('Venda', backref='usuario', lazy=True)
    clientes = db.relationship('Cliente', backref='usuario', lazy=True)
    fornecedores = db.relationship('Fornecedor', backref='usuario', lazy=True)
    compras = db.relationship('Compra', backref='usuario', lazy=True)
    produtos_auxiliares = db.relationship('ProdutoAuxiliar', backref='usuario', lazy=True)
    notas_fiscais = db.relationship('NotaFiscal', backref='usuario', lazy=True)
    tickets_suporte = db.relationship('TicketSuporte', backref='usuario', lazy=True)
    
    def set_password(self, password):
        """Define a senha do usuário usando hash seguro"""
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        """Verifica se a senha fornecida está correta"""
        return check_password_hash(self.password_hash, password)

class Produto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False, index=True)
    descricao = db.Column(db.Text)
    preco = db.Column(db.Float, nullable=False, index=True)
    estoque_atual = db.Column(db.Integer, default=0, index=True)
    estoque_minimo = db.Column(db.Integer, default=0)
    categoria = db.Column(db.String(50), index=True)
    codigo_barras = db.Column(db.String(50), unique=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    
    # Informações do Fornecedor
    fornecedor_nome = db.Column(db.String(100))
    fornecedor_cnpj = db.Column(db.String(18))
    fornecedor_contato = db.Column(db.String(100))
    fornecedor_telefone = db.Column(db.String(20))
    fornecedor_email = db.Column(db.String(120))
    fornecedor_endereco = db.Column(db.Text)
    preco_compra = db.Column(db.Float)
    prazo_entrega = db.Column(db.String(50))
    observacoes_fornecedor = db.Column(db.Text)
    
    # Relacionamentos
    itens_venda = db.relationship('ItemVenda', backref='produto', lazy=True)

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120))
    telefone = db.Column(db.String(20))
    endereco = db.Column(db.Text)
    cpf_cnpj = db.Column(db.String(20))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Campos para Cobrança
    forma_pagamento_preferida = db.Column(db.String(50))  # dinheiro, cartao, pix, boleto, transferencia
    limite_credito = db.Column(db.Float, default=0.0)
    dia_vencimento = db.Column(db.Integer)  # dia do mês para vencimento (1-31)
    
    # Campos para Notificações
    aceita_email = db.Column(db.Boolean, default=True)
    aceita_sms = db.Column(db.Boolean, default=False)
    aceita_whatsapp = db.Column(db.Boolean, default=True)
    horario_notificacoes = db.Column(db.String(20), default='09:00-18:00')  # horário preferido para notificações
    
    # Campos para Encomendas
    endereco_entrega = db.Column(db.Text)  # endereço específico para entregas
    instrucoes_entrega = db.Column(db.Text)  # instruções especiais para entrega
    pessoa_responsavel = db.Column(db.String(100))  # pessoa responsável por receber
    
    # Campos Comerciais
    tipo_cliente = db.Column(db.String(20), default='pessoa_fisica')  # pessoa_fisica, pessoa_juridica
    data_nascimento = db.Column(db.Date)
    status = db.Column(db.String(20), default='ativo')  # ativo, inativo, bloqueado
    observacoes = db.Column(db.Text)
    
    # Relacionamentos
    vendas = db.relationship('Venda', backref='cliente', lazy=True)

class Venda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_venda = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    valor_total = db.Column(db.Float, nullable=False, index=True)
    valor_desconto = db.Column(db.Float, default=0.0)  # Valor do desconto aplicado
    valor_final = db.Column(db.Float, nullable=False)  # Valor total após desconto
    status = db.Column(db.String(20), default='finalizada', index=True)  # finalizada, cancelada, pendente
    forma_pagamento = db.Column(db.String(30), index=True)  # dinheiro, cartao, pix, etc
    observacoes = db.Column(db.Text)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), index=True)
    cupom_id = db.Column(db.Integer, db.ForeignKey('cupom.id'), index=True)
    
    # Relacionamentos
    itens = db.relationship('ItemVenda', backref='venda', lazy=True, cascade='all, delete-orphan')

class ItemVenda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quantidade = db.Column(db.Integer, nullable=False)
    preco_unitario = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'), nullable=False)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'), nullable=False)

# Caixa: sessão de caixa diária
class CaixaSessao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_abertura = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    data_fechamento = db.Column(db.DateTime)
    status = db.Column(db.String(20), default='aberto', index=True)  # aberto, fechado
    saldo_inicial = db.Column(db.Float, default=0.0)
    saldo_fechamento = db.Column(db.Float)
    observacoes_abertura = db.Column(db.Text)
    observacoes_fechamento = db.Column(db.Text)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)

    movimentos = db.relationship('MovimentoCaixa', backref='sessao', lazy=True, cascade='all, delete-orphan')

class MovimentoCaixa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    tipo = db.Column(db.String(20), nullable=False, index=True)  # entrada, saida
    origem = db.Column(db.String(30), nullable=False, index=True)  # venda, suprimento, sangria, ajuste
    valor = db.Column(db.Float, nullable=False)
    descricao = db.Column(db.String(200))
    forma_pagamento = db.Column(db.String(30))  # quando origem for venda
    referencia_id = db.Column(db.Integer)  # id de venda, quando aplicável
    sessao_id = db.Column(db.Integer, db.ForeignKey('caixa_sessao.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, index=True)

# Novos modelos para controle de estoque
class Fornecedor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    razao_social = db.Column(db.String(100))
    cnpj = db.Column(db.String(18))
    email = db.Column(db.String(120))
    telefone = db.Column(db.String(20))
    endereco = db.Column(db.Text)
    cidade = db.Column(db.String(50))
    estado = db.Column(db.String(2))
    cep = db.Column(db.String(9))
    contato = db.Column(db.String(100))
    observacoes = db.Column(db.Text)
    status = db.Column(db.String(20), default='ativo')  # ativo, inativo
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Campos específicos para fornecedores
    
    # Informações Fiscais e Contratuais
    inscricao_estadual = db.Column(db.String(20))  # Inscrição Estadual
    inscricao_municipal = db.Column(db.String(20))  # Inscrição Municipal
    regime_tributario = db.Column(db.String(30))  # Simples, Presumido, Real
    categoria_fornecedor = db.Column(db.String(50))  # Matéria-prima, Serviços, Produtos acabados, etc.
    
    # Informações Financeiras
    limite_credito = db.Column(db.Float, default=0.0)  # Limite de crédito
    prazo_pagamento = db.Column(db.Integer, default=30)  # Prazo de pagamento em dias
    forma_pagamento_preferida = db.Column(db.String(50))  # PIX, Transferência, Boleto, etc.
    banco = db.Column(db.String(100))  # Nome do banco
    agencia = db.Column(db.String(20))  # Agência
    conta = db.Column(db.String(20))  # Conta corrente
    chave_pix = db.Column(db.String(100))  # Chave PIX
    
    # Informações de Contato Avançadas
    telefone_alternativo = db.Column(db.String(20))  # Telefone alternativo
    contato_financeiro = db.Column(db.String(100))  # Contato do financeiro
    contato_comercial = db.Column(db.String(100))  # Contato comercial
    contato_tecnico = db.Column(db.String(100))  # Contato técnico
    email_financeiro = db.Column(db.String(120))  # Email do financeiro
    email_comercial = db.Column(db.String(120))  # Email comercial
    email_tecnico = db.Column(db.String(120))  # Email técnico
    
    # Endereço Detalhado
    logradouro = db.Column(db.String(200))  # Rua, Avenida, etc.
    numero = db.Column(db.String(20))  # Número
    bairro = db.Column(db.String(100))  # Bairro
    complemento = db.Column(db.String(100))  # Complemento
    uf = db.Column(db.String(2))  # UF (padronizado)
    
    # Informações Operacionais
    tempo_entrega = db.Column(db.Integer)  # Tempo de entrega em dias
    modalidade_frete = db.Column(db.String(30))  # CIF, FOB, etc.
    condicoes_entrega = db.Column(db.Text)  # Condições especiais de entrega
    certificacoes = db.Column(db.Text)  # Certificações (ISO, etc.)
    
    # Avaliação e Controle
    nota_qualidade = db.Column(db.Float, default=0.0)  # Nota de qualidade (0-10)
    nota_pontualidade = db.Column(db.Float, default=0.0)  # Nota de pontualidade (0-10)
    nota_preco = db.Column(db.Float, default=0.0)  # Nota de preço (0-10)
    nota_atendimento = db.Column(db.Float, default=0.0)  # Nota de atendimento (0-10)
    data_ultima_avaliacao = db.Column(db.DateTime)  # Data da última avaliação
    
    # Relacionamentos
    compras = db.relationship('Compra', backref='fornecedor', lazy=True)

class Compra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero_compra = db.Column(db.String(50), unique=True)
    data_compra = db.Column(db.DateTime, default=datetime.utcnow)
    data_entrega = db.Column(db.DateTime)
    valor_total = db.Column(db.Float, nullable=False, default=0)
    status = db.Column(db.String(20), default='pendente')  # pendente, confirmada, entregue, cancelada
    forma_pagamento = db.Column(db.String(30))
    observacoes = db.Column(db.Text)
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('fornecedor.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Relacionamentos
    itens = db.relationship('ItemCompra', backref='compra', lazy=True, cascade='all, delete-orphan')
    notas_fiscais = db.relationship('NotaFiscal', backref='compra', lazy=True)

class ItemCompra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quantidade = db.Column(db.Integer, nullable=False)
    preco_unitario = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    compra_id = db.Column(db.Integer, db.ForeignKey('compra.id'), nullable=False)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'), nullable=False)

class ProdutoAuxiliar(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.Text)
    categoria = db.Column(db.String(50))
    unidade = db.Column(db.String(20))  # kg, litros, metros, etc
    preco_unitario = db.Column(db.Float, nullable=False)
    estoque_atual = db.Column(db.Float, default=0)
    estoque_minimo = db.Column(db.Float, default=0)
    codigo_interno = db.Column(db.String(50))
    observacoes = db.Column(db.Text)
    status = db.Column(db.String(20), default='ativo')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Informações do Fornecedor
    fornecedor_nome = db.Column(db.String(100))
    fornecedor_cnpj = db.Column(db.String(18))
    fornecedor_contato = db.Column(db.String(100))
    fornecedor_telefone = db.Column(db.String(20))
    fornecedor_email = db.Column(db.String(120))
    fornecedor_endereco = db.Column(db.Text)
    preco_compra = db.Column(db.Float)
    prazo_entrega = db.Column(db.String(50))
    observacoes_fornecedor = db.Column(db.Text)

class NotaFiscal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero_nf = db.Column(db.String(50), nullable=False)
    serie = db.Column(db.String(10))
    modelo = db.Column(db.String(10), default='NFe')  # NFe, NFCe, etc
    chave_acesso = db.Column(db.String(50))
    data_emissao = db.Column(db.DateTime, default=datetime.utcnow)
    data_saida = db.Column(db.DateTime)
    valor_total = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='pendente')  # pendente, emitida, cancelada
    tipo_operacao = db.Column(db.String(20))  # entrada, saida
    cliente_fornecedor = db.Column(db.String(100))
    observacoes = db.Column(db.Text)
    compra_id = db.Column(db.Integer, db.ForeignKey('compra.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

class TicketSuporte(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.Text, nullable=False)
    categoria = db.Column(db.String(50))  # bug, melhoria, duvida, etc
    prioridade = db.Column(db.String(20), default='media')  # baixa, media, alta, urgente
    status = db.Column(db.String(20), default='aberto')  # aberto, em_andamento, resolvido, fechado
    data_abertura = db.Column(db.DateTime, default=datetime.utcnow)
    data_fechamento = db.Column(db.DateTime)
    anexos = db.Column(db.Text)  # JSON com paths dos anexos
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

class Cupom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), nullable=False, unique=True, index=True)
    descricao = db.Column(db.String(200))
    tipo_desconto = db.Column(db.String(20), nullable=False)  # 'percentual' ou 'valor_fixo'
    valor_desconto = db.Column(db.Float, nullable=False)
    valor_minimo_compra = db.Column(db.Float, default=0.0)
    valor_maximo_desconto = db.Column(db.Float)  # Para desconto percentual
    limite_uso = db.Column(db.Integer)  # Número máximo de usos (None = ilimitado)
    usos_realizados = db.Column(db.Integer, default=0)
    data_inicio = db.Column(db.DateTime, nullable=True)  # Permitir NULL para cupons sem período
    data_fim = db.Column(db.DateTime, nullable=True)    # Permitir NULL para cupons sem período
    ativo = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Relacionamentos
    vendas = db.relationship('Venda', backref='cupom', lazy=True)

class RespostaTicket(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    mensagem = db.Column(db.Text, nullable=False)
    data_resposta = db.Column(db.DateTime, default=datetime.utcnow)
    is_admin = db.Column(db.Boolean, default=False)
    ticket_id = db.Column(db.Integer, db.ForeignKey('ticket_suporte.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

class UserSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    dark_mode = db.Column(db.Boolean, default=False)
    notifications = db.Column(db.Boolean, default=True)
    auto_logout = db.Column(db.Integer, default=30)  # minutes
    language = db.Column(db.String(5), default='pt')
    timezone = db.Column(db.String(50), default='America/Sao_Paulo')
    dashboard_refresh = db.Column(db.Integer, default=30)  # seconds
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ========== ADMIN DATA VIEWER ==========
@app.route('/admin')
@admin_required
@cache.memoize(timeout=300)  # Cache por 5 minutos
def admin_home():
    from datetime import datetime, timedelta
    
    # Estatísticas Gerais
    stats = {
        'total_usuarios': User.query.count(),
        'usuarios_ativos': User.query.filter(User.created_at >= datetime.now() - timedelta(days=30)).count(),
        'usuarios_admin': User.query.filter_by(role='admin').count(),
    }
    
    # Estatísticas de Clientes
    clientes_stats = {
        'total': Cliente.query.count(),
        'novos_este_mes': Cliente.query.filter(Cliente.created_at >= datetime.now().replace(day=1)).count(),
        'sem_vendas': db.session.query(Cliente).outerjoin(Venda).filter(Venda.id.is_(None)).count(),
    }
    
    # Estatísticas de Produtos
    produtos_stats = {
        'total': Produto.query.count(),
        'estoque_baixo': Produto.query.filter(Produto.estoque_atual <= Produto.estoque_minimo).count(),
        'sem_estoque': Produto.query.filter_by(estoque_atual=0).count(),
        'valor_total_estoque': db.session.query(db.func.sum(Produto.estoque_atual * Produto.preco)).scalar() or 0,
    }
    
    # Estatísticas de Vendas
    vendas_stats = {
        'total_vendas': Venda.query.count(),
        'vendas_este_mes': Venda.query.filter(
            Venda.data_venda >= datetime.now().replace(day=1),
            Venda.status == 'finalizada'
        ).count(),
        'faturamento_mes': db.session.query(db.func.sum(Venda.valor_total)).filter(
            Venda.data_venda >= datetime.now().replace(day=1),
            Venda.status == 'finalizada'
        ).scalar() or 0,
        'ticket_medio': db.session.query(db.func.avg(Venda.valor_total)).filter(
            Venda.status == 'finalizada'
        ).scalar() or 0,
    }
    
    # Estatísticas de Fornecedores
    fornecedores_stats = {
        'total': Fornecedor.query.count(),
        'ativos': Fornecedor.query.filter_by(status='ativo').count(),
        'inativos': Fornecedor.query.filter_by(status='inativo').count(),
    }
    
    # Estatísticas de Compras
    compras_stats = {
        'total': Compra.query.count(),
        'pendentes': Compra.query.filter_by(status='pendente').count(),
        'confirmadas': Compra.query.filter_by(status='confirmada').count(),
        'valor_total_pendente': db.session.query(db.func.sum(Compra.valor_total)).filter(
            Compra.status == 'pendente'
        ).scalar() or 0,
    }
    
    # Estatísticas de Produtos Auxiliares
    produtos_aux_stats = {
        'total': ProdutoAuxiliar.query.count(),
        'ativos': ProdutoAuxiliar.query.filter_by(status='ativo').count(),
        'estoque_baixo': ProdutoAuxiliar.query.filter(
            ProdutoAuxiliar.estoque_atual <= ProdutoAuxiliar.estoque_minimo
        ).count(),
    }
    
    # Estatísticas de Notas Fiscais
    notas_stats = {
        'total': NotaFiscal.query.count(),
        'pendentes': NotaFiscal.query.filter_by(status='pendente').count(),
        'emitidas': NotaFiscal.query.filter_by(status='emitida').count(),
        'valor_total': db.session.query(db.func.sum(NotaFiscal.valor_total)).scalar() or 0,
    }
    
    # Estatísticas de Suporte
    suporte_stats = {
        'tickets_abertos': TicketSuporte.query.filter(TicketSuporte.status.in_(['aberto', 'em_andamento'])).count(),
        'tickets_resolvidos': TicketSuporte.query.filter_by(status='resolvido').count(),
        'tickets_fechados': TicketSuporte.query.filter_by(status='fechado').count(),
        'total_respostas': RespostaTicket.query.count(),
    }
    
    # Dados para gráficos
    # Vendas dos últimos 7 dias
    vendas_7_dias = []
    for i in range(7):
        data = datetime.now().date() - timedelta(days=i)
        vendas_dia = Venda.query.filter(
            db.func.date(Venda.data_venda) == data,
            Venda.status == 'finalizada'
        ).count()
        vendas_7_dias.append({
            'data': data.strftime('%d/%m'),
            'vendas': vendas_dia
        })
    vendas_7_dias.reverse()
    
    # Top 5 produtos mais vendidos
    top_produtos = db.session.query(
        Produto.nome,
        db.func.sum(ItemVenda.quantidade).label('total_vendido')
    ).join(ItemVenda).join(Venda).filter(
        Venda.status == 'finalizada'
    ).group_by(Produto.id, Produto.nome).order_by(
        db.func.sum(ItemVenda.quantidade).desc()
    ).limit(5).all()
    
    # Usuários recentes
    usuarios_recentes = User.query.order_by(User.created_at.desc()).limit(5).all()
    
    # Tickets recentes
    tickets_recentes = TicketSuporte.query.order_by(TicketSuporte.data_abertura.desc()).limit(5).all()
    
    return render_template('admin/index.html', 
                         stats=stats,
                         clientes_stats=clientes_stats,
                         produtos_stats=produtos_stats,
                         vendas_stats=vendas_stats,
                         fornecedores_stats=fornecedores_stats,
                         compras_stats=compras_stats,
                         produtos_aux_stats=produtos_aux_stats,
                         notas_stats=notas_stats,
                         suporte_stats=suporte_stats,
                         vendas_7_dias=vendas_7_dias,
                         top_produtos=top_produtos,
                         usuarios_recentes=usuarios_recentes,
                         tickets_recentes=tickets_recentes)


MODEL_MAP = {
    'user': User,
    'cliente': Cliente,
    'produto': Produto,
    'venda': Venda,
    'fornecedor': Fornecedor,
    'compra': Compra,
    'produto_auxiliar': ProdutoAuxiliar,
    'nota_fiscal': NotaFiscal,
    'ticket_suporte': TicketSuporte,
    'resposta_ticket': RespostaTicket,
    'user_settings': UserSettings,
    'item_venda': ItemVenda,
    'item_compra': ItemCompra,
}

@app.route('/admin/table/<string:table_name>')
@admin_required
def admin_table(table_name: str):
    model = MODEL_MAP.get(table_name)
    if not model:
        flash('Tabela não reconhecida.', 'error')
        return redirect(url_for('admin_home'))
    
    # Parâmetros de paginação
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(per_page, 100)  # Limitar máximo de itens por página
    
    # Query com paginação
    pagination = model.query.paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    rows = pagination.items
    items = [serialize_model(r) for r in rows]
    columns = list(items[0].keys()) if items else [c.name for c in model.__table__.columns]
    
    return render_template('admin/table.html', 
                         table_name=table_name, 
                         columns=columns, 
                         items=items,
                         pagination=pagination)

@app.route('/admin/table/<string:table_name>/csv')
@admin_required
def admin_table_csv(table_name: str):
    import csv
    from io import StringIO
    model = MODEL_MAP.get(table_name)
    if not model:
        flash('Tabela não reconhecida.', 'error')
        return redirect(url_for('admin_home'))
    rows = model.query.all()
    items = [serialize_model(r) for r in rows]
    output = StringIO()
    if items:
        writer = csv.DictWriter(output, fieldnames=list(items[0].keys()))
        writer.writeheader()
        writer.writerows(items)
    else:
        writer = csv.writer(output)
        writer.writerow([c.name for c in model.__table__.columns])
    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = f'attachment; filename={table_name}.csv'
    return resp

# Rotas de Autenticação
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        return redirect(url_for('login'))
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('auth/login.html', google_configured=bool(google))

@app.route('/login', methods=['GET', 'POST'])
@csrf.exempt
@safe_rate_limit("5 per minute")  # Rate limiting para login
def login():
    if request.method == 'POST':
        # Sanitizar entrada
        email = sanitize_input(request.form.get('email', '').strip())
        password = request.form.get('password', '')
        
        # Validar email
        if not validate_email(email):
            flash('Formato de email inválido!', 'error')
            return render_template('auth/login.html', google_configured=bool(google))
        
        user = User.query.filter_by(email=email).first()
        
        # Verificar se conta está bloqueada
        if user and is_account_locked(user):
            flash('Conta temporariamente bloqueada. Tente novamente mais tarde.', 'error')
            return render_template('auth/login.html', google_configured=bool(google))
        
        if user and user.password_hash and check_password_hash(user.password_hash, password):
            # Verificar se 2FA está habilitado
            if user.two_factor_enabled:
                # Redirecionar para verificação 2FA
                session['temp_user_id'] = user.id
                session['2fa_required'] = True
                return redirect(url_for('verify_2fa'))
            else:
                # Login normal com verificação de dispositivo e sessão
                device_id = get_device_identifier(request)
                unlink_device = request.form.get('unlink_device') == 'true' or request.args.get('unlink_device') == 'true'
                is_admin_role = (user.role == 'admin')

                if user.device_fingerprint:
                    if user.device_fingerprint != device_id:
                        if is_admin_role and unlink_device:
                            user.device_fingerprint = device_id
                            user.device_last_updated = datetime.utcnow()
                        else:
                            msg = 'Login bloqueado: dispositivo diferente do vinculado.'
                            if is_admin_role:
                                msg += ' Você pode desvincular o dispositivo anterior para prosseguir.'
                            flash(msg, 'warning')
                            return render_template('auth/login.html', google_configured=bool(google))
                else:
                    # Primeiro login: vincula dispositivo
                    user.device_fingerprint = device_id
                    user.device_last_updated = datetime.utcnow()

                # Emitir token de sessão única
                sess_token = issue_session_token(user)

                # Carregar configurações do usuário
                settings = get_user_settings(user.id)

                # Resetar tentativas falhadas e atualizar último login
                reset_failed_login(user)
                user.last_login = datetime.utcnow()
                db.session.commit()

                # Setar sessão e cookie
                session['user_id'] = user.id
                session['username'] = user.username
                session['empresa'] = user.empresa
                session['role'] = user.role
                session['dark_mode'] = settings.dark_mode
                session['active_session_id'] = sess_token
                session.permanent = True

                flash('Login realizado com sucesso!', 'success')
                resp = redirect(url_for('dashboard'))
                set_device_cookie(resp, device_id)
                return resp
        else:
            # Incrementar tentativas falhadas
            if user:
                increment_failed_login(user)
            flash('Email ou senha incorretos!', 'error')
    
    return render_template('auth/login.html', google_configured=bool(google))

# ========== ROTAS API AUXILIARES ==========

@app.route('/api/cep/<string:cep>')
def api_cep(cep):
    """API para buscar dados do CEP"""
    try:
        data = get_cep_data(cep)
        if data:
            return jsonify({
                'success': True,
                'data': data
            })
        else:
            return jsonify({
                'success': False,
                'message': 'CEP não encontrado'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Erro ao buscar CEP'
        }), 500

# ========== ROTAS API MOBILE ==========

@app.route('/api/auth/login', methods=['POST'])
@csrf.exempt
@safe_rate_limit("10 per minute")
def api_login():
    """Login para aplicações mobile"""
    data = request.get_json()
    
    if not data or not data.get('email') or not data.get('password'):
        return jsonify({'error': 'Email e senha são obrigatórios'}), 400
    
    email = sanitize_input(data['email'].strip())
    password = data['password']
    
    # Validar email
    if not validate_email(email):
        return jsonify({'error': 'Formato de email inválido'}), 400
    
    user = User.query.filter_by(email=email).first()
    
    # Verificar se conta está bloqueada
    if user and is_account_locked(user):
        return jsonify({'error': 'Conta temporariamente bloqueada'}), 423
    
    if user and user.password_hash and check_password_hash(user.password_hash, password):
        # Verificar se 2FA está habilitado
        if user.two_factor_enabled:
            return jsonify({
                'error': '2FA required',
                'requires_2fa': True,
                'user_id': user.id
            }), 200
        # Verificar dispositivo
        device_id = get_device_identifier(request)
        is_admin_role = (user.role == 'admin')
        unlink_device = (data.get('unlink_device') in ['true', True]) if data else False
        if user.device_fingerprint:
            if user.device_fingerprint != device_id:
                if is_admin_role and unlink_device:
                    user.device_fingerprint = device_id
                    user.device_last_updated = datetime.utcnow()
                else:
                    return jsonify({'error': 'device_mismatch', 'message': 'Dispositivo diferente do vinculado', 'can_unlink': is_admin_role}), 403
        else:
            user.device_fingerprint = device_id
            user.device_last_updated = datetime.utcnow()

        # Criar tokens JWT
        access_token = create_access_token(identity=str(user.id))
        refresh_token = create_refresh_token(identity=str(user.id))
        
        # Atualizar último login e sessão
        reset_failed_login(user)
        user.last_login = datetime.utcnow()
        issue_session_token(user)
        db.session.commit()
        
        return jsonify({
            'access_token': access_token,
            'refresh_token': refresh_token,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'empresa': user.empresa,
                'role': user.role,
                'two_factor_enabled': user.two_factor_enabled
            }
        }), 200
    else:
        # Incrementar tentativas falhadas
        if user:
            increment_failed_login(user)
        return jsonify({'error': 'Email ou senha incorretos'}), 401

@app.route('/api/auth/unlink-device', methods=['POST'])
@csrf.exempt
@login_required
def api_unlink_device():
    """Admin atualiza o dispositivo vinculado para o dispositivo atual."""
    user = User.query.get(session.get('user_id'))
    if not user or user.role != 'admin':
        return jsonify({'success': False, 'message': 'Permissão negada'}), 403
    new_device_id = get_device_identifier(request)
    user.device_fingerprint = new_device_id
    user.device_last_updated = datetime.utcnow()
    issue_session_token(user)
    db.session.commit()
    session['active_session_id'] = user.active_session_id
    return jsonify({'success': True, 'message': 'Dispositivo atualizado com sucesso'}), 200

@app.route('/api/auth/refresh', methods=['POST'])
@csrf.exempt
@jwt_required(refresh=True)
def api_refresh():
    """Refresh token para aplicações mobile"""
    current_user_id = get_jwt_identity()
    user = User.query.get(current_user_id)
    
    if not user:
        return jsonify({'error': 'Usuário não encontrado'}), 404
    
    new_access_token = create_access_token(identity=str(current_user_id))
    
    return jsonify({
        'access_token': new_access_token,
        'user': {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'empresa': user.empresa,
            'role': user.role,
            'two_factor_enabled': user.two_factor_enabled
        }
    }), 200

@app.route('/api/auth/verify-2fa', methods=['POST'])
@csrf.exempt
@safe_rate_limit("10 per minute")
def api_verify_2fa():
    """Verificação 2FA para mobile"""
    data = request.get_json()
    
    if not data or not data.get('user_id'):
        return jsonify({'error': 'ID do usuário é obrigatório'}), 400
    
    user = User.query.get(data['user_id'])
    if not user:
        return jsonify({'error': 'Usuário não encontrado'}), 404
    
    token = data.get('token', '').strip()
    backup_code = data.get('backup_code', '').strip()
    
    # Verificar token 2FA ou código de backup
    if token:
        if verify_2fa_token(user.two_factor_secret, token):
            # Login bem-sucedido
            access_token = create_access_token(identity=str(user.id))
            refresh_token = create_refresh_token(identity=str(user.id))
            
            reset_failed_login(user)
            
            return jsonify({
                'access_token': access_token,
                'refresh_token': refresh_token,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'empresa': user.empresa,
                    'role': user.role,
                    'two_factor_enabled': user.two_factor_enabled
                }
            }), 200
        else:
            return jsonify({'error': 'Código 2FA inválido'}), 401
    
    elif backup_code:
        # Verificar código de backup
        import json
        backup_codes = json.loads(user.backup_codes or '[]')
        if backup_code.upper() in backup_codes:
            # Remover código usado
            backup_codes.remove(backup_code.upper())
            user.backup_codes = json.dumps(backup_codes)
            db.session.commit()
            
            # Login bem-sucedido
            access_token = create_access_token(identity=str(user.id))
            refresh_token = create_refresh_token(identity=str(user.id))
            
            reset_failed_login(user)
            
            return jsonify({
                'access_token': access_token,
                'refresh_token': refresh_token,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'empresa': user.empresa,
                    'role': user.role,
                    'two_factor_enabled': user.two_factor_enabled
                }
            }), 200
        else:
            return jsonify({'error': 'Código de backup inválido'}), 401
    
    else:
        return jsonify({'error': 'Digite um código 2FA ou código de backup'}), 400

@app.route('/api/auth/logout', methods=['POST'])
@csrf.exempt
@jwt_required()
def api_logout():
    """Logout para aplicações mobile"""
    # Em uma implementação mais robusta, você adicionaria o token a uma blacklist
    return jsonify({'message': 'Logout realizado com sucesso'}), 200

# ========== ROTAS 2FA ==========

@app.route('/verify-2fa', methods=['GET', 'POST'])
@csrf.exempt
@safe_rate_limit("10 per minute")
def verify_2fa():
    if not session.get('2fa_required'):
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        token = request.form.get('token', '').strip()
        backup_code = request.form.get('backup_code', '').strip()
        
        user = User.query.get(session['temp_user_id'])
        
        if not user:
            flash('Sessão expirada. Faça login novamente.', 'error')
            return redirect(url_for('login'))
        
        # Verificar token 2FA ou código de backup
        if token:
            if verify_2fa_token(user.two_factor_secret, token):
                # Login bem-sucedido
                session.pop('temp_user_id', None)
                session.pop('2fa_required', None)
                session['user_id'] = user.id
                session['username'] = user.username
                session['empresa'] = user.empresa
                session['role'] = user.role
                session.permanent = True
                
                # Carregar configurações do usuário
                settings = get_user_settings(user.id)
                session['dark_mode'] = settings.dark_mode
                
                reset_failed_login(user)
                flash('Login realizado com sucesso!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Código 2FA inválido!', 'error')
        elif backup_code:
            # Verificar código de backup
            import json
            backup_codes = json.loads(user.backup_codes or '[]')
            if backup_code.upper() in backup_codes:
                # Remover código usado
                backup_codes.remove(backup_code.upper())
                user.backup_codes = json.dumps(backup_codes)
                db.session.commit()
                
                # Login bem-sucedido
                session.pop('temp_user_id', None)
                session.pop('2fa_required', None)
                session['user_id'] = user.id
                session['username'] = user.username
                session['empresa'] = user.empresa
                session['role'] = user.role
                session.permanent = True
                
                settings = get_user_settings(user.id)
                session['dark_mode'] = settings.dark_mode
                
                reset_failed_login(user)
                flash('Login realizado com código de backup!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Código de backup inválido!', 'error')
        else:
            flash('Digite um código 2FA ou código de backup!', 'error')
    
    return render_template('auth/verify_2fa.html')

@app.route('/setup-2fa')
@login_required
@admin_required
def setup_2fa():
    user = User.query.get(session['user_id'])
    
    if user.two_factor_enabled:
        flash('2FA já está habilitado para sua conta!', 'info')
        return redirect(url_for('dashboard'))
    
    # Gerar nova chave secreta
    secret = generate_2fa_secret()
    qr_code = generate_2fa_qr_code(user.email, secret)
    
    # Salvar temporariamente no banco (será confirmado após verificação)
    user.two_factor_secret = secret
    db.session.commit()
    
    return render_template('auth/setup_2fa.html', qr_code=qr_code, secret=secret)

@app.route('/confirm-2fa', methods=['POST'])
@csrf.exempt
@login_required
@admin_required
@safe_rate_limit("5 per minute")
def confirm_2fa():
    user = User.query.get(session['user_id'])
    token = request.form.get('token', '').strip()
    
    if not token:
        flash('Digite o código 2FA!', 'error')
        return redirect(url_for('setup_2fa'))
    
    if verify_2fa_token(user.two_factor_secret, token):
        # Habilitar 2FA
        user.two_factor_enabled = True
        user.backup_codes = generate_backup_codes()
        db.session.commit()
        
        flash('2FA habilitado com sucesso! Guarde seus códigos de backup.', 'success')
        return redirect(url_for('dashboard'))
    else:
        flash('Código 2FA inválido!', 'error')
        return redirect(url_for('setup_2fa'))

@app.route('/disable-2fa', methods=['POST'])
@csrf.exempt
@login_required
@admin_required
@safe_rate_limit("3 per minute")
def disable_2fa():
    user = User.query.get(session['user_id'])
    token = request.form.get('token', '').strip()
    
    if not verify_2fa_token(user.two_factor_secret, token):
        flash('Código 2FA inválido!', 'error')
        return redirect(url_for('dashboard'))
    
    # Desabilitar 2FA
    user.two_factor_enabled = False
    user.two_factor_secret = None
    user.backup_codes = None
    db.session.commit()
    
    flash('2FA desabilitado com sucesso!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/register', methods=['GET', 'POST'])
@csrf.exempt
@safe_rate_limit("3 per minute")  # Rate limiting para registro
def register():
    if request.method == 'POST':
        # Sanitizar entrada
        username = sanitize_input(request.form.get('username', '').strip())
        email = sanitize_input(request.form.get('email', '').strip())
        password = request.form.get('password', '')
        empresa = sanitize_input(request.form.get('empresa', '').strip())
        
        # Validações de entrada
        if not username or len(username) < 3:
            flash('Nome de usuário deve ter pelo menos 3 caracteres!', 'error')
            return render_template('auth/register.html', google_configured=bool(google))
        
        if not validate_email(email):
            flash('Formato de email inválido!', 'error')
            return render_template('auth/register.html', google_configured=bool(google))
        
        # Validar força da senha
        is_valid, message = validate_password_strength(password)
        if not is_valid:
            flash(message, 'error')
            return render_template('auth/register.html', google_configured=bool(google))
        
        if not empresa or len(empresa) < 2:
            flash('Nome da empresa deve ter pelo menos 2 caracteres!', 'error')
            return render_template('auth/register.html', google_configured=bool(google))
        
        # Verificar se usuário já existe por username
        if User.query.filter_by(username=username).first():
            flash('Nome de usuário já cadastrado!', 'error')
            return render_template('auth/register.html', google_configured=bool(google))
        
        # Verificar se usuário já existe por email
        if User.query.filter_by(email=email).first():
            flash('Email já cadastrado!', 'error')
            return render_template('auth/register.html', google_configured=bool(google))
        
        try:
            # Criar novo usuário
            user = User(
                username=username,
                email=email,
                password_hash=generate_password_hash(password),
                empresa=empresa
            )
            
            db.session.add(user)
            db.session.commit()
            
            flash('Conta criada com sucesso! Faça login para continuar.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            db.session.rollback()
            flash('Erro ao criar conta. Tente novamente.', 'error')
            return render_template('auth/register.html', google_configured=bool(google))
    
    return render_template('auth/register.html', google_configured=bool(google))

# Rotas de Autenticação Google
@app.route('/login/google')
def google_login():
    if not google:
        flash('Autenticação Google não configurada. Configure as credenciais no arquivo .env', 'error')
        return redirect(url_for('login'))
    return google.authorize(callback=url_for('google_authorized', _external=True))

@app.route('/login/google/authorized')
def google_authorized():
    if not google:
        flash('Autenticação Google não configurada.', 'error')
        return redirect(url_for('login'))
    
    resp = google.authorized_response()
    if resp is None:
        flash('Acesso negado pelo Google.', 'error')
        return redirect(url_for('login'))
    
    session['google_token'] = (resp['access_token'], '')
    user_info = google.get('userinfo')
    
    if user_info.status != 200:
        flash('Erro ao obter informações do Google.', 'error')
        return redirect(url_for('login'))
    
    google_data = user_info.data
    google_id = google_data['id']
    email = google_data['email']
    name = google_data['name']
    picture = google_data.get('picture', '')
    
    # Verificar se usuário já existe
    user = User.query.filter_by(google_id=google_id).first()
    
    if not user:
        # Verificar se email já existe (usuário pode ter se registrado manualmente)
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            # Vincular conta Google ao usuário existente
            existing_user.google_id = google_id
            existing_user.avatar_url = picture
            db.session.commit()
            user = existing_user
        else:
            # Criar novo usuário
            username = email.split('@')[0]  # Usar parte do email como username
            # Garantir username único
            counter = 1
            original_username = username
            while User.query.filter_by(username=username).first():
                username = f"{original_username}{counter}"
                counter += 1
            
            user = User(
                username=username,
                email=email,
                empresa='Empresa Google',  # Valor padrão
                google_id=google_id,
                avatar_url=picture,
                password_hash=None  # Sem senha para usuários Google
            )
            db.session.add(user)
            db.session.commit()
    
    # Fazer login do usuário
    session['user_id'] = user.id
    session['username'] = user.username
    session['empresa'] = user.empresa
    session['role'] = user.role
    session['google_avatar'] = user.avatar_url
    
    flash(f'Bem-vindo, {user.username}!', 'success')
    return redirect(url_for('dashboard'))

def get_google_oauth_token():
    return session.get('google_token')

if google:
    google.tokengetter(get_google_oauth_token)

@app.route('/logout')
def logout():
    session.clear()
    flash('Logout realizado com sucesso!', 'info')
    return redirect(url_for('login'))

# Dashboard
@app.route('/dashboard')
@login_required
def dashboard():
    user_id = session['user_id']
    
    # Estatísticas básicas
    total_produtos = Produto.query.filter_by(user_id=user_id).count()
    total_clientes = Cliente.query.filter_by(user_id=user_id).count()
    total_fornecedores = Fornecedor.query.filter_by(user_id=user_id, status='ativo').count()
    total_produtos_auxiliares = ProdutoAuxiliar.query.filter_by(user_id=user_id, status='ativo').count()
    
    # Vendas do mês atual
    inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    vendas_mes = Venda.query.filter(
        Venda.user_id == user_id,
        Venda.data_venda >= inicio_mes,
        Venda.status == 'finalizada'
    ).all()
    
    total_vendas_mes = sum(venda.valor_total for venda in vendas_mes)
    
    # Compras do mês atual
    compras_mes = Compra.query.filter(
        Compra.user_id == user_id,
        Compra.data_compra >= inicio_mes
    ).all()
    
    total_compras_mes = sum(compra.valor_total for compra in compras_mes)
    
    # Produtos com estoque baixo
    produtos_estoque_baixo = Produto.query.filter(
        Produto.user_id == user_id,
        Produto.estoque_atual <= Produto.estoque_minimo
    ).count()
    
    # Produtos auxiliares com estoque baixo
    produtos_auxiliares_estoque_baixo = ProdutoAuxiliar.query.filter(
        ProdutoAuxiliar.user_id == user_id,
        ProdutoAuxiliar.estoque_atual <= ProdutoAuxiliar.estoque_minimo
    ).count()
    
    # Tickets de suporte abertos
    tickets_abertos = TicketSuporte.query.filter(
        TicketSuporte.user_id == user_id,
        TicketSuporte.status.in_(['aberto', 'em_andamento'])
    ).count()
    
    # Vendas recentes
    vendas_recentes = Venda.query.filter_by(user_id=user_id).order_by(Venda.data_venda.desc()).limit(5).all()
    
    # Gráfico de vendas dos últimos 7 dias
    vendas_7_dias = []
    for i in range(7):
        data = datetime.now().date() - timedelta(days=i)
        vendas_dia = Venda.query.filter(
            Venda.user_id == user_id,
            db.func.date(Venda.data_venda) == data,
            Venda.status == 'finalizada'
        ).count()
        vendas_7_dias.append({'data': data.strftime('%d/%m'), 'vendas': vendas_dia})
    
    vendas_7_dias.reverse()
    
    return render_template('dashboard.html',
                         total_produtos=total_produtos,
                         total_clientes=total_clientes,
                         total_fornecedores=total_fornecedores,
                         total_produtos_auxiliares=total_produtos_auxiliares,
                         total_vendas_mes=total_vendas_mes,
                         total_compras_mes=total_compras_mes,
                         produtos_estoque_baixo=produtos_estoque_baixo,
                         produtos_auxiliares_estoque_baixo=produtos_auxiliares_estoque_baixo,
                         tickets_abertos=tickets_abertos,
                         vendas_recentes=vendas_recentes,
                         vendas_7_dias=vendas_7_dias)

# Produtos
@app.route('/produtos')
@login_required
def produtos():
    if is_admin():
        produtos = Produto.query.order_by(Produto.nome).all()
    else:
        user_id = session['user_id']
        produtos = Produto.query.filter_by(user_id=user_id).order_by(Produto.nome).all()
    return render_template('produtos/list.html', produtos=produtos)

@app.route('/produtos/novo', methods=['GET', 'POST'])
@login_required
def novo_produto():
    if request.method == 'POST':
        # Convert price from Brazilian format (comma) to float
        preco_str = request.form['preco'].replace(',', '.')
        
        # Processar preço de compra se fornecido
        preco_compra_str = request.form.get('preco_compra', '').replace(',', '.')
        preco_compra = float(preco_compra_str) if preco_compra_str else None
        
        produto = Produto(
            nome=request.form['nome'],
            descricao=request.form['descricao'],
            preco=float(preco_str),
            estoque_atual=int(request.form['estoque_atual']),
            estoque_minimo=int(request.form['estoque_minimo']),
            categoria=request.form['categoria'],
            codigo_barras=request.form['codigo_barras'],
            user_id=session['user_id'],
            # Campos de fornecedor
            fornecedor_nome=request.form.get('fornecedor_nome'),
            fornecedor_cnpj=request.form.get('fornecedor_cnpj'),
            fornecedor_contato=request.form.get('fornecedor_contato'),
            fornecedor_telefone=request.form.get('fornecedor_telefone'),
            fornecedor_email=request.form.get('fornecedor_email'),
            fornecedor_endereco=request.form.get('fornecedor_endereco'),
            preco_compra=preco_compra,
            prazo_entrega=request.form.get('prazo_entrega'),
            observacoes_fornecedor=request.form.get('observacoes_fornecedor')
        )
        
        db.session.add(produto)
        db.session.commit()
        
        flash('Produto cadastrado com sucesso!', 'success')
        return redirect(url_for('produtos'))
    
    # Buscar categorias existentes para sugestões
    user_id = session['user_id']
    categorias_existentes = db.session.query(Produto.categoria).filter(
        Produto.user_id == user_id,
        Produto.categoria.isnot(None),
        Produto.categoria != ''
    ).distinct().all()
    
    categorias = [cat[0] for cat in categorias_existentes]
    
    return render_template('produtos/form.html', categorias_existentes=categorias)

@app.route('/produtos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_produto(id):
    produto = Produto.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    if request.method == 'POST':
        # Convert price from Brazilian format (comma) to float
        preco_str = request.form['preco'].replace(',', '.')
        
        # Processar preço de compra se fornecido
        preco_compra_str = request.form.get('preco_compra', '').replace(',', '.')
        preco_compra = float(preco_compra_str) if preco_compra_str else None
        
        produto.nome = request.form['nome']
        produto.descricao = request.form['descricao']
        produto.preco = float(preco_str)
        produto.estoque_atual = int(request.form['estoque_atual'])
        produto.estoque_minimo = int(request.form['estoque_minimo'])
        produto.categoria = request.form['categoria']
        produto.codigo_barras = request.form['codigo_barras']
        
        # Campos de fornecedor
        produto.fornecedor_nome = request.form.get('fornecedor_nome')
        produto.fornecedor_cnpj = request.form.get('fornecedor_cnpj')
        produto.fornecedor_contato = request.form.get('fornecedor_contato')
        produto.fornecedor_telefone = request.form.get('fornecedor_telefone')
        produto.fornecedor_email = request.form.get('fornecedor_email')
        produto.fornecedor_endereco = request.form.get('fornecedor_endereco')
        produto.preco_compra = preco_compra
        produto.prazo_entrega = request.form.get('prazo_entrega')
        produto.observacoes_fornecedor = request.form.get('observacoes_fornecedor')
        
        db.session.commit()
        flash('Produto atualizado com sucesso!', 'success')
        return redirect(url_for('produtos'))
    
    # Buscar categorias existentes para sugestões
    user_id = session['user_id']
    categorias_existentes = db.session.query(Produto.categoria).filter(
        Produto.user_id == user_id,
        Produto.categoria.isnot(None),
        Produto.categoria != ''
    ).distinct().all()
    
    categorias = [cat[0] for cat in categorias_existentes]
    
    return render_template('produtos/form.html', produto=produto, categorias_existentes=categorias)

@app.route('/produtos/excluir/<int:id>')
@login_required
def excluir_produto(id):
    produto = Produto.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    db.session.delete(produto)
    db.session.commit()
    flash('Produto excluído com sucesso!', 'success')
    return redirect(url_for('produtos'))

# Clientes
@app.route('/clientes')
@login_required
def clientes():
    if is_admin():
        clientes = Cliente.query.order_by(Cliente.nome).all()
    else:
        user_id = session['user_id']
        clientes = Cliente.query.filter_by(user_id=user_id).order_by(Cliente.nome).all()
    return render_template('clientes/list.html', clientes=clientes)

@app.route('/clientes/novo', methods=['GET', 'POST'])
@login_required
def novo_cliente():
    if request.method == 'POST':
        # Validações
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        telefone = request.form.get('telefone', '').strip()
        tipo_cliente = request.form.get('tipo_cliente', 'pessoa_fisica')
        
        # Validar nome obrigatório
        if not nome:
            flash('Nome é obrigatório!', 'error')
            return render_template('clientes/form.html')
        
        # Validar email se fornecido
        if email and not validate_email(email):
            flash('Email inválido!', 'error')
            return render_template('clientes/form.html')
        
        # Validar CPF/CNPJ se fornecido
        if cpf_cnpj:
            if tipo_cliente == 'pessoa_fisica':
                if not validate_cpf(cpf_cnpj):
                    flash('CPF inválido!', 'error')
                    return render_template('clientes/form.html')
            else:
                if not validate_cnpj(cpf_cnpj):
                    flash('CNPJ inválido!', 'error')
                    return render_template('clientes/form.html')
        
        # Verificar se email já existe
        if email:
            existing_cliente = Cliente.query.filter_by(email=email, user_id=session['user_id']).first()
            if existing_cliente:
                flash('Este email já está cadastrado!', 'error')
                return render_template('clientes/form.html')
        
        # Verificar se CPF/CNPJ já existe
        if cpf_cnpj:
            existing_cliente = Cliente.query.filter_by(cpf_cnpj=cpf_cnpj, user_id=session['user_id']).first()
            if existing_cliente:
                flash('Este CPF/CNPJ já está cadastrado!', 'error')
                return render_template('clientes/form.html')
        
        # Converter limite de crédito
        limite_credito_str = request.form.get('limite_credito', '0,00').replace(',', '.')
        try:
            limite_credito = float(limite_credito_str)
        except ValueError:
            limite_credito = 0.0
        
        # Converter dia de vencimento
        dia_vencimento = request.form.get('dia_vencimento')
        dia_vencimento = int(dia_vencimento) if dia_vencimento else None
        
        # Converter data de nascimento
        data_nascimento_str = request.form.get('data_nascimento')
        data_nascimento = None
        if data_nascimento_str:
            try:
                data_nascimento = datetime.strptime(data_nascimento_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Criar cliente
        cliente = Cliente(
            nome=nome,
            email=email if email else None,
            telefone=format_phone(telefone) if telefone else None,
            endereco=request.form.get('endereco'),
            cpf_cnpj=format_cpf_cnpj(cpf_cnpj) if cpf_cnpj else None,
            tipo_cliente=tipo_cliente,
            data_nascimento=data_nascimento,
            status=request.form.get('status', 'ativo'),
            
            # Cobrança
            forma_pagamento_preferida=request.form.get('forma_pagamento_preferida'),
            limite_credito=limite_credito,
            dia_vencimento=dia_vencimento,
            
            # Notificações
            aceita_email=bool(request.form.get('aceita_email')),
            aceita_sms=bool(request.form.get('aceita_sms')),
            aceita_whatsapp=bool(request.form.get('aceita_whatsapp')),
            horario_notificacoes=request.form.get('horario_notificacoes', '09:00-18:00'),
            
            # Entrega
            endereco_entrega=request.form.get('endereco_entrega'),
            instrucoes_entrega=request.form.get('instrucoes_entrega'),
            pessoa_responsavel=request.form.get('pessoa_responsavel'),
            
            # Comercial
            observacoes=request.form.get('observacoes'),
            user_id=session['user_id']
        )
        
        db.session.add(cliente)
        db.session.commit()
        
        flash('Cliente cadastrado com sucesso!', 'success')
        return redirect(url_for('clientes'))
    
    return render_template('clientes/form.html')

@app.route('/clientes/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(id):
    try:
        # Lógica de acesso: admin pode editar qualquer cliente, usuário comum apenas os seus
        if is_admin():
            cliente = Cliente.query.filter_by(id=id).first()
        else:
            cliente = Cliente.query.filter_by(id=id, user_id=session['user_id']).first()
        
        if not cliente:
            flash(f'Cliente com ID {id} não encontrado ou você não tem permissão para editá-lo.', 'error')
            return redirect(url_for('clientes'))
        
    except Exception as e:
        flash(f'Erro ao carregar cliente: {str(e)}', 'error')
        return redirect(url_for('clientes'))
    
    if request.method == 'POST':
        # Validações
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        telefone = request.form.get('telefone', '').strip()
        tipo_cliente = request.form.get('tipo_cliente', 'pessoa_fisica')
        
        # Validar nome obrigatório
        if not nome:
            flash('Nome é obrigatório!', 'error')
            return render_template('clientes/form.html', cliente=cliente)
        
        # Validar email se fornecido
        if email and not validate_email(email):
            flash('Email inválido!', 'error')
            return render_template('clientes/form.html', cliente=cliente)
        
        # Validar CPF/CNPJ se fornecido
        if cpf_cnpj:
            if tipo_cliente == 'pessoa_fisica':
                if not validate_cpf(cpf_cnpj):
                    flash('CPF inválido!', 'error')
                    return render_template('clientes/form.html', cliente=cliente)
            else:
                if not validate_cnpj(cpf_cnpj):
                    flash('CNPJ inválido!', 'error')
                    return render_template('clientes/form.html', cliente=cliente)
        
        # Verificar se email já existe (exceto o próprio cliente)
        if email:
            existing_cliente = Cliente.query.filter_by(email=email, user_id=session['user_id']).filter(Cliente.id != id).first()
            if existing_cliente:
                flash('Este email já está cadastrado!', 'error')
                return render_template('clientes/form.html', cliente=cliente)
        
        # Verificar se CPF/CNPJ já existe (exceto o próprio cliente)
        if cpf_cnpj:
            existing_cliente = Cliente.query.filter_by(cpf_cnpj=cpf_cnpj, user_id=session['user_id']).filter(Cliente.id != id).first()
            if existing_cliente:
                flash('Este CPF/CNPJ já está cadastrado!', 'error')
                return render_template('clientes/form.html', cliente=cliente)
        
        # Converter limite de crédito
        limite_credito_str = request.form.get('limite_credito', '0,00').replace(',', '.')
        try:
            limite_credito = float(limite_credito_str)
        except ValueError:
            limite_credito = 0.0
        
        # Converter dia de vencimento
        dia_vencimento = request.form.get('dia_vencimento')
        dia_vencimento = int(dia_vencimento) if dia_vencimento else None
        
        # Converter data de nascimento
        data_nascimento_str = request.form.get('data_nascimento')
        data_nascimento = None
        if data_nascimento_str:
            try:
                data_nascimento = datetime.strptime(data_nascimento_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Atualizar cliente
        cliente.nome = nome
        cliente.email = email if email else None
        cliente.telefone = format_phone(telefone) if telefone else None
        cliente.cpf_cnpj = format_cpf_cnpj(cpf_cnpj) if cpf_cnpj else None
        cliente.tipo_cliente = tipo_cliente
        cliente.data_nascimento = data_nascimento
        cliente.status = request.form.get('status', 'ativo')
        
        # Endereço
        cliente.cep = request.form.get('cep')
        cliente.logradouro = request.form.get('logradouro')
        cliente.numero = request.form.get('numero')
        cliente.bairro = request.form.get('bairro')
        cliente.cidade = request.form.get('cidade')
        cliente.uf = request.form.get('uf')
        
        # Cobrança
        cliente.forma_pagamento_preferida = request.form.get('forma_pagamento_preferida')
        cliente.limite_credito = limite_credito
        cliente.dia_vencimento = dia_vencimento
        
        # Notificações
        cliente.aceita_email = bool(request.form.get('aceita_email'))
        cliente.aceita_sms = bool(request.form.get('aceita_sms'))
        cliente.aceita_whatsapp = bool(request.form.get('aceita_whatsapp'))
        cliente.horario_notificacoes = request.form.get('horario_notificacoes', '09:00-18:00')
        
        # Entrega
        cliente.endereco_entrega = request.form.get('endereco_entrega')
        cliente.instrucoes_entrega = request.form.get('instrucoes_entrega')
        cliente.pessoa_responsavel = request.form.get('pessoa_responsavel')
        
        # Comercial
        cliente.observacoes = request.form.get('observacoes')
        
        db.session.commit()
        flash('Cliente atualizado com sucesso!', 'success')
        return redirect(url_for('clientes'))
    
    return render_template('clientes/form.html', cliente=cliente)

@app.route('/clientes/excluir/<int:id>')
@login_required
def excluir_cliente(id):
    # Lógica de acesso: admin pode excluir qualquer cliente, usuário comum apenas os seus
    if is_admin():
        cliente = Cliente.query.filter_by(id=id).first()
    else:
        cliente = Cliente.query.filter_by(id=id, user_id=session['user_id']).first()
    
    if not cliente:
        flash(f'Cliente com ID {id} não encontrado ou você não tem permissão para excluí-lo.', 'error')
        return redirect(url_for('clientes'))
    
    db.session.delete(cliente)
    db.session.commit()
    flash('Cliente excluído com sucesso!', 'success')
    return redirect(url_for('clientes'))

# Vendas
@app.route('/vendas')
@login_required
def vendas():
    if is_admin():
        vendas = Venda.query.order_by(Venda.data_venda.desc()).all()
    else:
        user_id = session['user_id']
        vendas = Venda.query.filter_by(user_id=user_id).order_by(Venda.data_venda.desc()).all()
    return render_template('vendas/list.html', vendas=vendas)

@app.route('/vendas/nova', methods=['GET', 'POST'])
@login_required
def nova_venda():
    if request.method == 'POST':
        # Criar venda
        venda = Venda(
            cliente_id=int(request.form['cliente_id']) if request.form['cliente_id'] else None,
            valor_total=0,  # Será calculado
            valor_final=0,  # Será calculado
            forma_pagamento=request.form['forma_pagamento'],
            observacoes=request.form['observacoes'],
            user_id=session['user_id']
        )
        
        db.session.add(venda)
        db.session.flush()  # Para obter o ID da venda
        
        # Processar itens da venda
        total_venda = 0
        for key, value in request.form.items():
            if key.startswith('produto_') and value:
                produto_id = key.split('_')[1]
                quantidade = int(value)
                
                produto = Produto.query.filter_by(id=produto_id, user_id=session['user_id']).first()
                if produto and produto.estoque_atual >= quantidade:
                    # Criar item da venda
                    item = ItemVenda(
                        venda_id=venda.id,
                        produto_id=produto.id,
                        quantidade=quantidade,
                        preco_unitario=produto.preco,
                        subtotal=quantidade * produto.preco
                    )
                    
                    db.session.add(item)
                    
                    # Atualizar estoque
                    produto.estoque_atual -= quantidade
                    total_venda += item.subtotal
        
        # Processar cupom se fornecido
        cupom_id = None
        valor_desconto = 0
        valor_final = total_venda
        
        if request.form.get('cupom_codigo'):
            from datetime import datetime
            codigo_cupom = request.form['cupom_codigo'].upper().strip()
            cupom = Cupom.query.filter_by(codigo=codigo_cupom, user_id=session['user_id']).first()
            
            if cupom and cupom.ativo:
                agora = datetime.utcnow()
                if agora >= cupom.data_inicio and agora <= cupom.data_fim:
                    if not cupom.limite_uso or cupom.usos_realizados < cupom.limite_uso:
                        # Verificar valor mínimo
                        if total_venda >= cupom.valor_minimo_compra:
                            # Calcular desconto
                            if cupom.tipo_desconto == 'percentual':
                                valor_desconto = total_venda * (cupom.valor_desconto / 100)
                                if cupom.valor_maximo_desconto:
                                    valor_desconto = min(valor_desconto, cupom.valor_maximo_desconto)
                            else:
                                valor_desconto = min(cupom.valor_desconto, total_venda)
                            
                            valor_final = total_venda - valor_desconto
                            cupom_id = cupom.id
                            
                            # Atualizar uso do cupom
                            cupom.usos_realizados += 1
        
        # Atualizar valores da venda
        venda.valor_total = total_venda
        venda.valor_desconto = valor_desconto
        venda.valor_final = valor_final
        venda.cupom_id = cupom_id
        
        db.session.commit()
        flash('Venda realizada com sucesso!', 'success')
        return redirect(url_for('vendas'))
    
    user_id = session['user_id']
    produtos = Produto.query.filter_by(user_id=user_id).all()
    clientes = Cliente.query.filter_by(user_id=user_id).all()
    
    return render_template('vendas/form.html', produtos=produtos, clientes=clientes)

@app.route('/vendas/<int:id>')
@login_required
def detalhes_venda(id):
    venda = Venda.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    return render_template('vendas/detalhes.html', venda=venda)

# Rotas de Fornecedores
@app.route('/fornecedores')
@login_required
def fornecedores():
    if is_admin():
        fornecedores = Fornecedor.query.order_by(Fornecedor.nome).all()
    else:
        user_id = session['user_id']
        fornecedores = Fornecedor.query.filter_by(user_id=user_id).order_by(Fornecedor.nome).all()
    return render_template('fornecedores/list.html', fornecedores=fornecedores)

@app.route('/fornecedores/novo', methods=['GET', 'POST'])
@login_required
def novo_fornecedor():
    if request.method == 'POST':
        # Validações
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        telefone = request.form.get('telefone', '').strip()
        
        # Validar nome obrigatório
        if not nome:
            flash('Nome é obrigatório!', 'error')
            return render_template('fornecedores/form.html')
        
        # Validar email se fornecido
        if email and not validate_email(email):
            flash('Email inválido!', 'error')
            return render_template('fornecedores/form.html')
        
        # Validar CNPJ se fornecido
        if cnpj and not validate_cnpj(cnpj):
            flash('CNPJ inválido!', 'error')
            return render_template('fornecedores/form.html')
        
        # Verificar se email já existe
        if email:
            existing_fornecedor = Fornecedor.query.filter_by(email=email, user_id=session['user_id']).first()
            if existing_fornecedor:
                flash('Este email já está cadastrado!', 'error')
                return render_template('fornecedores/form.html')
        
        # Verificar se CNPJ já existe
        if cnpj:
            existing_fornecedor = Fornecedor.query.filter_by(cnpj=cnpj, user_id=session['user_id']).first()
            if existing_fornecedor:
                flash('Este CNPJ já está cadastrado!', 'error')
                return render_template('fornecedores/form.html')
        
        # Converter limite de crédito
        limite_credito_str = request.form.get('limite_credito', '0,00').replace(',', '.')
        try:
            limite_credito = float(limite_credito_str)
        except ValueError:
            limite_credito = 0.0
        
        # Converter prazo de pagamento
        prazo_pagamento = request.form.get('prazo_pagamento')
        prazo_pagamento = int(prazo_pagamento) if prazo_pagamento else 30
        
        # Criar fornecedor
        fornecedor = Fornecedor(
            nome=nome,
            razao_social=request.form.get('razao_social'),
            cnpj=format_cpf_cnpj(cnpj) if cnpj else None,
            email=email if email else None,
            telefone=format_phone(telefone) if telefone else None,
            status=request.form.get('status', 'ativo'),
            
            # Endereço
            cep=request.form.get('cep'),
            logradouro=request.form.get('logradouro'),
            numero=request.form.get('numero'),
            bairro=request.form.get('bairro'),
            cidade=request.form.get('cidade'),
            uf=request.form.get('uf'),
            complemento=request.form.get('complemento'),
            
            # Informações fiscais
            inscricao_estadual=request.form.get('inscricao_estadual'),
            inscricao_municipal=request.form.get('inscricao_municipal'),
            regime_tributario=request.form.get('regime_tributario'),
            categoria_fornecedor=request.form.get('categoria_fornecedor'),
            
            # Informações financeiras
            limite_credito=limite_credito,
            prazo_pagamento=prazo_pagamento,
            forma_pagamento_preferida=request.form.get('forma_pagamento_preferida'),
            banco=request.form.get('banco'),
            agencia=request.form.get('agencia'),
            conta=request.form.get('conta'),
            chave_pix=request.form.get('chave_pix'),
            
            # Contatos avançados
            contato_financeiro=request.form.get('contato_financeiro'),
            contato_comercial=request.form.get('contato_comercial'),
            email_financeiro=request.form.get('email_financeiro'),
            email_comercial=request.form.get('email_comercial'),
            
            # Observações
            observacoes=request.form.get('observacoes'),
            user_id=session['user_id']
        )
        
        db.session.add(fornecedor)
        db.session.commit()
        
        flash('Fornecedor cadastrado com sucesso!', 'success')
        return redirect(url_for('fornecedores'))
    
    return render_template('fornecedores/form.html')

@app.route('/fornecedores/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_fornecedor(id):
    fornecedor = Fornecedor.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    if request.method == 'POST':
        # Validações
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        telefone = request.form.get('telefone', '').strip()
        
        # Validar nome obrigatório
        if not nome:
            flash('Nome é obrigatório!', 'error')
            return render_template('fornecedores/form.html', fornecedor=fornecedor)
        
        # Validar email se fornecido
        if email and not validate_email(email):
            flash('Email inválido!', 'error')
            return render_template('fornecedores/form.html', fornecedor=fornecedor)
        
        # Validar CNPJ se fornecido
        if cnpj and not validate_cnpj(cnpj):
            flash('CNPJ inválido!', 'error')
            return render_template('fornecedores/form.html', fornecedor=fornecedor)
        
        # Verificar se email já existe (exceto o próprio fornecedor)
        if email:
            existing_fornecedor = Fornecedor.query.filter_by(email=email, user_id=session['user_id']).filter(Fornecedor.id != id).first()
            if existing_fornecedor:
                flash('Este email já está cadastrado!', 'error')
                return render_template('fornecedores/form.html', fornecedor=fornecedor)
        
        # Verificar se CNPJ já existe (exceto o próprio fornecedor)
        if cnpj:
            existing_fornecedor = Fornecedor.query.filter_by(cnpj=cnpj, user_id=session['user_id']).filter(Fornecedor.id != id).first()
            if existing_fornecedor:
                flash('Este CNPJ já está cadastrado!', 'error')
                return render_template('fornecedores/form.html', fornecedor=fornecedor)
        
        # Converter limite de crédito
        limite_credito_str = request.form.get('limite_credito', '0,00').replace(',', '.')
        try:
            limite_credito = float(limite_credito_str)
        except ValueError:
            limite_credito = 0.0
        
        # Converter prazo de pagamento
        prazo_pagamento = request.form.get('prazo_pagamento')
        prazo_pagamento = int(prazo_pagamento) if prazo_pagamento else 30
        
        # Atualizar fornecedor
        fornecedor.nome = nome
        fornecedor.razao_social = request.form.get('razao_social')
        fornecedor.cnpj = format_cpf_cnpj(cnpj) if cnpj else None
        fornecedor.email = email if email else None
        fornecedor.telefone = format_phone(telefone) if telefone else None
        fornecedor.status = request.form.get('status', 'ativo')
        
        # Endereço
        fornecedor.cep = request.form.get('cep')
        fornecedor.logradouro = request.form.get('logradouro')
        fornecedor.numero = request.form.get('numero')
        fornecedor.bairro = request.form.get('bairro')
        fornecedor.cidade = request.form.get('cidade')
        fornecedor.uf = request.form.get('uf')
        fornecedor.complemento = request.form.get('complemento')
        
        # Informações fiscais
        fornecedor.inscricao_estadual = request.form.get('inscricao_estadual')
        fornecedor.inscricao_municipal = request.form.get('inscricao_municipal')
        fornecedor.regime_tributario = request.form.get('regime_tributario')
        fornecedor.categoria_fornecedor = request.form.get('categoria_fornecedor')
        
        # Informações financeiras
        fornecedor.limite_credito = limite_credito
        fornecedor.prazo_pagamento = prazo_pagamento
        fornecedor.forma_pagamento_preferida = request.form.get('forma_pagamento_preferida')
        fornecedor.banco = request.form.get('banco')
        fornecedor.agencia = request.form.get('agencia')
        fornecedor.conta = request.form.get('conta')
        fornecedor.chave_pix = request.form.get('chave_pix')
        
        # Contatos avançados
        fornecedor.contato_financeiro = request.form.get('contato_financeiro')
        fornecedor.contato_comercial = request.form.get('contato_comercial')
        fornecedor.email_financeiro = request.form.get('email_financeiro')
        fornecedor.email_comercial = request.form.get('email_comercial')
        
        
        # Observações
        fornecedor.observacoes = request.form.get('observacoes')
        
        db.session.commit()
        flash('Fornecedor atualizado com sucesso!', 'success')
        return redirect(url_for('fornecedores'))
    
    return render_template('fornecedores/form.html', fornecedor=fornecedor)

@app.route('/fornecedores/excluir/<int:id>')
@login_required
def excluir_fornecedor(id):
    fornecedor = Fornecedor.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    db.session.delete(fornecedor)
    db.session.commit()
    flash('Fornecedor excluído com sucesso!', 'success')
    return redirect(url_for('fornecedores'))

# Rotas de Compras
@app.route('/compras')
@login_required
def compras():
    if is_admin():
        compras = Compra.query.order_by(Compra.data_compra.desc()).all()
    else:
        user_id = session['user_id']
        compras = Compra.query.filter_by(user_id=user_id).order_by(Compra.data_compra.desc()).all()
    return render_template('compras/list.html', compras=compras)

@app.route('/compras/nova', methods=['GET', 'POST'])
@login_required
def nova_compra():
    if request.method == 'POST':
        # Gerar número da compra
        ultima_compra = Compra.query.filter_by(user_id=session['user_id']).order_by(Compra.id.desc()).first()
        numero_compra = f"COMP{datetime.now().strftime('%Y%m%d')}{(ultima_compra.id + 1 if ultima_compra else 1):04d}"
        
        # Criar compra
        compra = Compra(
            numero_compra=numero_compra,
            fornecedor_id=int(request.form['fornecedor_id']),
            data_entrega=datetime.strptime(request.form['data_entrega'], '%Y-%m-%d') if request.form.get('data_entrega') else None,
            forma_pagamento=request.form.get('forma_pagamento'),
            observacoes=request.form.get('observacoes'),
            user_id=session['user_id']
        )
        
        db.session.add(compra)
        db.session.flush()
        
        # Processar itens da compra
        total_compra = 0
        for key, value in request.form.items():
            if key.startswith('produto_') and value:
                produto_id = key.split('_')[1]
                quantidade = int(value)
                
                produto = Produto.query.filter_by(id=produto_id, user_id=session['user_id']).first()
                if produto:
                    # Criar item da compra
                    item = ItemCompra(
                        compra_id=compra.id,
                        produto_id=produto.id,
                        quantidade=quantidade,
                        preco_unitario=produto.preco,
                        subtotal=quantidade * produto.preco
                    )
                    
                    db.session.add(item)
                    total_compra += item.subtotal
        
        # Atualizar valor total da compra
        compra.valor_total = total_compra
        
        db.session.commit()
        flash('Compra cadastrada com sucesso!', 'success')
        return redirect(url_for('compras'))
    
    user_id = session['user_id']
    # Não carregar produtos aqui - serão carregados via AJAX baseado no fornecedor selecionado
    fornecedores = Fornecedor.query.filter_by(user_id=user_id, status='ativo').all()
    
    return render_template('compras/form.html', produtos=[], fornecedores=fornecedores)

@app.route('/api/produtos-por-fornecedor/<int:fornecedor_id>')
@login_required
def api_produtos_por_fornecedor(fornecedor_id):
    """API para buscar produtos de um fornecedor específico"""
    try:
        # Buscar o fornecedor
        fornecedor = Fornecedor.query.filter_by(id=fornecedor_id, user_id=session['user_id']).first()
        if not fornecedor:
            return jsonify({'success': False, 'error': 'Fornecedor não encontrado'}), 404
        
        # Buscar produtos que têm informações deste fornecedor
        produtos = Produto.query.filter(
            Produto.user_id == session['user_id'],
            Produto.fornecedor_nome.ilike(f'%{fornecedor.nome}%')
        ).all()
        
        # Se não encontrar por nome, buscar por CNPJ se disponível
        if not produtos and fornecedor.cnpj:
            produtos = Produto.query.filter(
                Produto.user_id == session['user_id'],
                Produto.fornecedor_cnpj == fornecedor.cnpj
            ).all()
        
        return jsonify({
            'success': True,
            'data': [
                {
                    'id': produto.id,
                    'nome': produto.nome,
                    'preco': float(produto.preco),
                    'estoque_atual': produto.estoque_atual,
                    'categoria': produto.categoria,
                    'fornecedor_nome': produto.fornecedor_nome,
                    'preco_compra': float(produto.preco_compra) if produto.preco_compra else None
                } for produto in produtos
            ]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/compras/<int:id>')
@login_required
def detalhes_compra(id):
    compra = Compra.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    return render_template('compras/detalhes.html', compra=compra)

@app.route('/compras/confirmar/<int:id>')
@login_required
def confirmar_compra(id):
    compra = Compra.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    if compra.status == 'pendente':
        # Atualizar estoque dos produtos
        for item in compra.itens:
            item.produto.estoque_atual += item.quantidade
        
        compra.status = 'confirmada'
        db.session.commit()
        flash('Compra confirmada e estoque atualizado!', 'success')
    
    return redirect(url_for('compras'))

# Rotas de Produtos Auxiliares
@app.route('/produtos-auxiliares')
@login_required
def produtos_auxiliares():
    user_id = session['user_id']
    produtos = ProdutoAuxiliar.query.filter_by(user_id=user_id).order_by(ProdutoAuxiliar.nome).all()
    return render_template('produtos_auxiliares/list.html', produtos=produtos)

@app.route('/produtos-auxiliares/novo', methods=['GET', 'POST'])
@login_required
def novo_produto_auxiliar():
    if request.method == 'POST':
        try:
            # Debug: imprimir dados do formulário
            print("Dados do formulário:", request.form.to_dict())
            
            # Validar campos obrigatórios
            if not request.form.get('nome'):
                flash('Nome do produto é obrigatório!', 'error')
                return render_template('produtos_auxiliares/form.html')
            
            if not request.form.get('preco_unitario'):
                flash('Preço unitário é obrigatório!', 'error')
                return render_template('produtos_auxiliares/form.html')
            
            # Converter valores com tratamento de erro
            preco_unitario = 0.0
            estoque_atual = 0.0
            estoque_minimo = 0.0
            
            try:
                preco_unitario = float(request.form['preco_unitario'])
            except (ValueError, TypeError):
                preco_unitario = 0.0
            
            try:
                estoque_atual = float(request.form.get('estoque_atual', 0))
            except (ValueError, TypeError):
                estoque_atual = 0.0
            
            try:
                estoque_minimo = float(request.form.get('estoque_minimo', 0))
            except (ValueError, TypeError):
                estoque_minimo = 0.0
            
            # Processar preço de compra se fornecido
            preco_compra_str = request.form.get('preco_compra', '').replace(',', '.')
            preco_compra = float(preco_compra_str) if preco_compra_str else None
            
            produto = ProdutoAuxiliar(
                nome=request.form['nome'].strip(),
                descricao=request.form.get('descricao', '').strip(),
                categoria=request.form.get('categoria', '').strip(),
                unidade=request.form.get('unidade', '').strip(),
                preco_unitario=preco_unitario,
                estoque_atual=estoque_atual,
                estoque_minimo=estoque_minimo,
                codigo_interno=request.form.get('codigo_interno', '').strip(),
                observacoes=request.form.get('observacoes', '').strip(),
                user_id=session['user_id'],
                # Campos de fornecedor
                fornecedor_nome=request.form.get('fornecedor_nome'),
                fornecedor_cnpj=request.form.get('fornecedor_cnpj'),
                fornecedor_contato=request.form.get('fornecedor_contato'),
                fornecedor_telefone=request.form.get('fornecedor_telefone'),
                fornecedor_email=request.form.get('fornecedor_email'),
                fornecedor_endereco=request.form.get('fornecedor_endereco'),
                preco_compra=preco_compra,
                prazo_entrega=request.form.get('prazo_entrega'),
                observacoes_fornecedor=request.form.get('observacoes_fornecedor')
            )
            
            print("Produto criado:", produto.nome, produto.preco_unitario)
            
            db.session.add(produto)
            db.session.commit()
            
            flash('Produto auxiliar cadastrado com sucesso!', 'success')
            return redirect(url_for('produtos_auxiliares'))
            
        except Exception as e:
            db.session.rollback()
            print(f"Erro ao cadastrar produto auxiliar: {str(e)}")
            flash(f'Erro ao cadastrar produto auxiliar: {str(e)}', 'error')
            return render_template('produtos_auxiliares/form.html')
    
    return render_template('produtos_auxiliares/form.html')

@app.route('/produtos-auxiliares/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_produto_auxiliar(id):
    produto = ProdutoAuxiliar.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    if request.method == 'POST':
        # Processar preço de compra se fornecido
        preco_compra_str = request.form.get('preco_compra', '').replace(',', '.')
        preco_compra = float(preco_compra_str) if preco_compra_str else None
        
        produto.nome = request.form['nome']
        produto.descricao = request.form.get('descricao')
        produto.categoria = request.form.get('categoria')
        produto.unidade = request.form.get('unidade')
        produto.preco_unitario = float(request.form['preco_unitario'])
        produto.estoque_atual = float(request.form['estoque_atual'])
        produto.estoque_minimo = float(request.form['estoque_minimo'])
        produto.codigo_interno = request.form.get('codigo_interno')
        produto.observacoes = request.form.get('observacoes')
        
        # Campos de fornecedor
        produto.fornecedor_nome = request.form.get('fornecedor_nome')
        produto.fornecedor_cnpj = request.form.get('fornecedor_cnpj')
        produto.fornecedor_contato = request.form.get('fornecedor_contato')
        produto.fornecedor_telefone = request.form.get('fornecedor_telefone')
        produto.fornecedor_email = request.form.get('fornecedor_email')
        produto.fornecedor_endereco = request.form.get('fornecedor_endereco')
        produto.preco_compra = preco_compra
        produto.prazo_entrega = request.form.get('prazo_entrega')
        produto.observacoes_fornecedor = request.form.get('observacoes_fornecedor')
        
        db.session.commit()
        flash('Produto auxiliar atualizado com sucesso!', 'success')
        return redirect(url_for('produtos_auxiliares'))
    
    return render_template('produtos_auxiliares/form.html', produto=produto)

@app.route('/produtos-auxiliares/excluir/<int:id>')
@login_required
def excluir_produto_auxiliar(id):
    produto = ProdutoAuxiliar.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    db.session.delete(produto)
    db.session.commit()
    flash('Produto auxiliar excluído com sucesso!', 'success')
    return redirect(url_for('produtos_auxiliares'))

# Rotas de Nota Fiscal
@app.route('/notas-fiscais')
@login_required
def notas_fiscais():
    if is_admin():
        notas = NotaFiscal.query.order_by(NotaFiscal.data_emissao.desc()).all()
    else:
        user_id = session['user_id']
        notas = NotaFiscal.query.filter_by(user_id=user_id).order_by(NotaFiscal.data_emissao.desc()).all()
    return render_template('notas_fiscais/list.html', notas=notas)

@app.route('/notas-fiscais/nova', methods=['GET', 'POST'])
@login_required
def nova_nota_fiscal():
    if request.method == 'POST':
        nota = NotaFiscal(
            numero_nf=request.form['numero_nf'],
            serie=request.form.get('serie'),
            modelo=request.form.get('modelo', 'NFe'),
            chave_acesso=request.form.get('chave_acesso'),
            data_saida=datetime.strptime(request.form['data_saida'], '%Y-%m-%d') if request.form.get('data_saida') else None,
            valor_total=float(request.form['valor_total']),
            tipo_operacao=request.form.get('tipo_operacao'),
            cliente_fornecedor=request.form.get('cliente_fornecedor'),
            observacoes=request.form.get('observacoes'),
            compra_id=int(request.form['compra_id']) if request.form.get('compra_id') else None,
            user_id=session['user_id']
        )
        
        db.session.add(nota)
        db.session.commit()
        
        flash('Nota fiscal cadastrada com sucesso!', 'success')
        return redirect(url_for('notas_fiscais'))
    
    user_id = session['user_id']
    compras = Compra.query.filter_by(user_id=user_id).all()
    
    return render_template('notas_fiscais/form.html', compras=compras)

@app.route('/notas-fiscais/<int:id>')
@login_required
def detalhes_nota_fiscal(id):
    nota = NotaFiscal.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    return render_template('notas_fiscais/detalhes.html', nota=nota)

# Rotas de Suporte
@app.route('/suporte')
@login_required
def suporte():
    if is_admin():
        tickets = TicketSuporte.query.order_by(TicketSuporte.data_abertura.desc()).all()
    else:
        user_id = session['user_id']
        tickets = TicketSuporte.query.filter_by(user_id=user_id).order_by(TicketSuporte.data_abertura.desc()).all()
    return render_template('suporte/list.html', tickets=tickets)

@app.route('/suporte/novo', methods=['GET', 'POST'])
@login_required
def novo_ticket():
    if request.method == 'POST':
        ticket = TicketSuporte(
            titulo=request.form['titulo'],
            descricao=request.form['descricao'],
            categoria=request.form.get('categoria'),
            prioridade=request.form.get('prioridade', 'media'),
            user_id=session['user_id']
        )
        
        db.session.add(ticket)
        db.session.commit()
        
        flash('Ticket de suporte criado com sucesso!', 'success')
        return redirect(url_for('suporte'))
    
    return render_template('suporte/form.html')

@app.route('/suporte/<int:id>')
@login_required
def detalhes_ticket(id):
    ticket = TicketSuporte.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    return render_template('suporte/detalhes.html', ticket=ticket)

@app.route('/suporte/<int:id>/responder', methods=['POST'])
@login_required
def responder_ticket(id):
    ticket = TicketSuporte.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    resposta = RespostaTicket(
        mensagem=request.form['mensagem'],
        ticket_id=ticket.id,
        user_id=session['user_id']
    )
    
    db.session.add(resposta)
    
    # Atualizar status do ticket se necessário
    if ticket.status == 'resolvido':
        ticket.status = 'aberto'
    
    db.session.commit()
    flash('Resposta enviada com sucesso!', 'success')
    return redirect(url_for('detalhes_ticket', id=id))

# API para buscar produtos
# ====== CAIXA (CaixaSessao) ======

def get_sessao_caixa_aberta(user_id):
    return CaixaSessao.query.filter_by(user_id=user_id, status='aberto')\
        .order_by(CaixaSessao.data_abertura.desc()).first()

@app.route('/caixa')
@login_required
def caixa_dashboard():
    user_id = session['user_id']
    sessao = get_sessao_caixa_aberta(user_id)

    # Período de cálculo: sessão aberta ou dia atual
    inicio = (sessao.data_abertura if sessao else datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0))
    fim = datetime.utcnow()

    vendas = Venda.query.filter(
        Venda.user_id == user_id,
        Venda.data_venda.between(inicio, fim),
        Venda.status == 'finalizada'
    ).all()

    total_vendas = sum(v.valor_total for v in vendas)
    qtd_vendas = len(vendas)
    ticket_medio = (total_vendas / qtd_vendas) if qtd_vendas else 0

    # Quebra por forma de pagamento
    formas = ['dinheiro', 'cartao', 'pix', 'boleto', 'transferencia']
    por_pagamento = {fp: sum(v.valor_total for v in vendas if (v.forma_pagamento or '').lower() == fp) for fp in formas}

    # Movimentos do caixa (se houver sessão)
    movimentos = []
    if sessao:
        movimentos = MovimentoCaixa.query.filter_by(sessao_id=sessao.id).order_by(MovimentoCaixa.created_at.desc()).limit(10).all()

    return render_template('caixa/index.html', sessao=sessao, inicio=inicio, fim=fim,
                           total_vendas=total_vendas, qtd_vendas=qtd_vendas, ticket_medio=ticket_medio,
                           por_pagamento=por_pagamento, movimentos=movimentos)

@app.route('/caixa/abrir', methods=['POST'])
@login_required
def abrir_caixa():
    user_id = session['user_id']
    if get_sessao_caixa_aberta(user_id):
        flash('Já existe um caixa aberto.', 'warning')
        return redirect(url_for('caixa_dashboard'))

    saldo_raw = request.form.get('saldo_inicial')
    if saldo_raw is None or str(saldo_raw).strip() == '':
        flash('Informe o saldo inicial para abrir o caixa.', 'warning')
        return redirect(url_for('caixa_dashboard'))
    try:
        saldo_inicial = float(str(saldo_raw).replace(',', '.'))
    except ValueError:
        flash('Saldo inicial inválido.', 'danger')
        return redirect(url_for('caixa_dashboard'))
    if saldo_inicial < 0:
        flash('Saldo inicial não pode ser negativo.', 'warning')
        return redirect(url_for('caixa_dashboard'))

    observacoes = request.form.get('observacoes_abertura') or None

    sessao = CaixaSessao(
        data_abertura=datetime.utcnow(),
        status='aberto',
        saldo_inicial=saldo_inicial,
        observacoes_abertura=observacoes,
        user_id=user_id
    )
    db.session.add(sessao)
    db.session.commit()
    flash('Caixa aberto com sucesso!', 'success')
    return redirect(url_for('caixa_dashboard'))

@app.route('/caixa/fechar', methods=['POST'])
@login_required
def fechar_caixa():
    user_id = session['user_id']
    sessao = get_sessao_caixa_aberta(user_id)
    if not sessao:
        flash('Nenhum caixa aberto.', 'warning')
        return redirect(url_for('caixa_dashboard'))

    saldo_raw = request.form.get('saldo_fechamento')
    if saldo_raw is None or str(saldo_raw).strip() == '':
        flash('Informe o saldo no fechamento.', 'warning')
        return redirect(url_for('caixa_dashboard'))
    try:
        saldo_fechamento = float(str(saldo_raw).replace(',', '.'))
    except ValueError:
        flash('Saldo de fechamento inválido.', 'danger')
        return redirect(url_for('caixa_dashboard'))
    if saldo_fechamento < 0:
        flash('Saldo de fechamento não pode ser negativo.', 'warning')
        return redirect(url_for('caixa_dashboard'))

    observacoes = request.form.get('observacoes_fechamento') or None

    # Conferência de caixa: saldo esperado vs informado
    movimentos = MovimentoCaixa.query.filter_by(sessao_id=sessao.id).all()
    entradas = sum(m.valor for m in movimentos if m.tipo == 'entrada')
    saidas = sum(m.valor for m in movimentos if m.tipo == 'saida')
    esperado = (sessao.saldo_inicial or 0) + entradas - saidas
    diferenca = saldo_fechamento - esperado
    if abs(diferenca) > 0.01:
        flash(f'Diferença no fechamento de R$ {diferenca:.2f}. Confira o caixa.', 'danger')
        return redirect(url_for('caixa_dashboard'))

    sessao.data_fechamento = datetime.utcnow()
    sessao.status = 'fechado'
    sessao.saldo_fechamento = saldo_fechamento
    sessao.observacoes_fechamento = observacoes
    db.session.commit()
    flash('Caixa fechado com sucesso!', 'success')
    return redirect(url_for('caixa_dashboard'))

@app.route('/caixa/venda', methods=['POST'])
@login_required
def caixa_registrar_venda():
    user_id = session['user_id']
    sessao = get_sessao_caixa_aberta(user_id)
    if not sessao:
        flash('Nenhum caixa aberto.', 'warning')
        return redirect(url_for('caixa_dashboard'))

    forma_pagamento = (request.form.get('forma_pagamento') or '').lower().strip()
    observacoes = request.form.get('observacoes') or None

    # Dados do produto (se fornecidos)
    produto_id = request.form.get('produto_id')
    quantidade_raw = request.form.get('quantidade')

    # Caminho com produto selecionado: calcula total, cria ItemVenda e atualiza estoque
    if produto_id and quantidade_raw:
        try:
            produto_id = int(produto_id)
            quantidade = int(quantidade_raw)
        except ValueError:
            flash('Produto ou quantidade inválidos.', 'danger')
            return redirect(url_for('caixa_dashboard'))

        if quantidade <= 0:
            flash('Informe uma quantidade válida (>= 1).', 'warning')
            return redirect(url_for('caixa_dashboard'))

        produto = Produto.query.filter_by(id=produto_id, user_id=user_id).first()
        if not produto:
            flash('Produto não encontrado.', 'danger')
            return redirect(url_for('caixa_dashboard'))

        if produto.estoque_atual < quantidade:
            flash('Quantidade solicitada acima do estoque disponível.', 'warning')
            return redirect(url_for('caixa_dashboard'))

        if not forma_pagamento:
            flash('Informe a forma de pagamento.', 'warning')
            return redirect(url_for('caixa_dashboard'))

        valor = float(produto.preco) * quantidade

        venda = Venda(
            data_venda=datetime.utcnow(),
            valor_total=valor,
            valor_desconto=0.0,
            valor_final=valor,
            status='finalizada',
            forma_pagamento=forma_pagamento,
            observacoes=observacoes,
            user_id=user_id
        )
        db.session.add(venda)
        db.session.flush()  # obter venda.id

        item = ItemVenda(
            quantidade=quantidade,
            preco_unitario=float(produto.preco),
            subtotal=valor,
            venda_id=venda.id,
            produto_id=produto.id
        )
        db.session.add(item)

        # Atualiza estoque
        produto.estoque_atual = produto.estoque_atual - quantidade
        db.session.add(produto)

        movimento = MovimentoCaixa(
            tipo='entrada',
            origem='venda',
            valor=valor,
            descricao=f'Venda rápida: {produto.nome} x{quantidade}',
            forma_pagamento=forma_pagamento,
            referencia_id=venda.id,
            sessao_id=sessao.id,
            user_id=user_id
        )
        db.session.add(movimento)
        db.session.commit()

        flash('Venda registrada com sucesso no caixa!', 'success')
        return redirect(url_for('caixa_dashboard'))

    # Caminho sem produto: manter registro rápido por valor informado
    try:
        valor = float(request.form.get('valor', 0) or 0)
    except ValueError:
        flash('Valor inválido.', 'danger')
        return redirect(url_for('caixa_dashboard'))

    if valor <= 0 or not forma_pagamento:
        flash('Informe valor e forma de pagamento.', 'warning')
        return redirect(url_for('caixa_dashboard'))

    venda = Venda(
        data_venda=datetime.utcnow(),
        valor_total=valor,
        valor_desconto=0.0,
        valor_final=valor,
        status='finalizada',
        forma_pagamento=forma_pagamento,
        observacoes=observacoes,
        user_id=user_id
    )
    db.session.add(venda)
    db.session.flush()  # obter venda.id

    movimento = MovimentoCaixa(
        tipo='entrada',
        origem='venda',
        valor=valor,
        descricao='Venda rápida registrada no caixa',
        forma_pagamento=forma_pagamento,
        referencia_id=venda.id,
        sessao_id=sessao.id,
        user_id=user_id
    )
    db.session.add(movimento)
    db.session.commit()

    flash('Venda registrada com sucesso no caixa!', 'success')
    return redirect(url_for('caixa_dashboard'))

@app.route('/api/caixa/dashboard')
@csrf.exempt
@login_required
def api_caixa_dashboard():
    user_id = session['user_id']
    sessao = get_sessao_caixa_aberta(user_id)
    inicio = (sessao.data_abertura if sessao else datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0))
    fim = datetime.utcnow()

    vendas = Venda.query.filter(
        Venda.user_id == user_id,
        Venda.data_venda.between(inicio, fim),
        Venda.status == 'finalizada'
    ).all()
    total_vendas = sum(v.valor_total for v in vendas)
    qtd_vendas = len(vendas)
    ticket_medio = (total_vendas / qtd_vendas) if qtd_vendas else 0
    formas = ['dinheiro', 'cartao', 'pix', 'boleto', 'transferencia']
    por_pagamento = {fp: sum(v.valor_total for v in vendas if (v.forma_pagamento or '').lower() == fp) for fp in formas}

    return jsonify({'success': True,
                    'sessao_aberta': bool(sessao),
                    'inicio': inicio.isoformat(),
                    'fim': fim.isoformat(),
                    'total_vendas': total_vendas,
                    'qtd_vendas': qtd_vendas,
                    'ticket_medio': ticket_medio,
                    'por_pagamento': por_pagamento})

# Listagem de sessões de caixa
@app.route('/caixa/sessoes')
@login_required
def caixa_sessoes():
    user_id = session['user_id']
    status = request.args.get('status')  # aberto|fechado|None
    query = CaixaSessao.query.filter_by(user_id=user_id)
    if status in ['aberto', 'fechado']:
        query = query.filter(CaixaSessao.status == status)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    paginated = query.order_by(CaixaSessao.data_abertura.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template('caixa/sessoes.html', paginated=paginated, status=status)

# Detalhes de uma sessão específica
@app.route('/caixa/sessao/<int:sessao_id>')
@login_required
def caixa_sessao_detalhes(sessao_id):
    user_id = session['user_id']
    sessao = CaixaSessao.query.filter_by(id=sessao_id, user_id=user_id).first_or_404()
    movimentos = MovimentoCaixa.query.filter_by(sessao_id=sessao.id).order_by(MovimentoCaixa.created_at.desc()).all()
    entradas = sum(m.valor for m in movimentos if m.tipo == 'entrada')
    saidas = sum(m.valor for m in movimentos if m.tipo == 'saida')
    saldo_parcial = (sessao.saldo_inicial or 0) + entradas - saidas
    return render_template('caixa/detalhes.html', sessao=sessao, movimentos=movimentos, entradas=entradas, saidas=saidas, saldo_parcial=saldo_parcial)

# Importação de movimentos via CSV
@app.route('/caixa/sessao/<int:sessao_id>/importar', methods=['POST'])
@login_required
def importar_movimentos_caixa(sessao_id):
    user_id = session['user_id']
    sessao = CaixaSessao.query.filter_by(id=sessao_id, user_id=user_id).first_or_404()
    file = request.files.get('arquivo')
    if not file:
        flash('Nenhum arquivo enviado.', 'warning')
        return redirect(url_for('caixa_sessao_detalhes', sessao_id=sessao.id))
    try:
        # Ler CSV usando encoding seguro
        stream = file.stream.read().decode('utf-8', errors='ignore')
        reader = csv.DictReader(stream.splitlines())
        headers = reader.fieldnames or []
        rows = []
        max_rows = 1000
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append({k: (row.get(k) or '').strip() for k in headers})
        if not rows:
            flash('CSV vazio ou inválido.', 'warning')
            return redirect(url_for('caixa_sessao_detalhes', sessao_id=sessao.id))
        default_map = {
            'tipo': next((h for h in headers if h.lower() == 'tipo'), None),
            'origem': next((h for h in headers if h.lower() == 'origem'), None),
            'valor': next((h for h in headers if h.lower() in ['valor', 'amount']), None),
            'descricao': next((h for h in headers if h.lower() == 'descricao'), None),
            'forma_pagamento': next((h for h in headers if h.lower() in ['forma_pagamento', 'pagamento', 'payment_method']), None),
            'referencia_id': next((h for h in headers if h.lower() in ['referencia_id', 'ref_id', 'id']), None),
            'data': next((h for h in headers if h.lower() in ['data', 'date', 'created_at']), None)
        }
        session['csv_import_caixa'] = {
            'sessao_id': sessao.id,
            'headers': headers,
            'rows': rows,
            'default_map': default_map
        }
        flash(f'Pré-visualização carregada ({min(len(rows), max_rows)} linhas).', 'info')
        return render_template('caixa/import_preview.html', sessao=sessao, headers=headers, rows=rows[:50], default_map=default_map, total_rows=len(rows))
    except Exception as e:
        db.session.rollback()
        flash(f'Falha ao importar: {e}', 'danger')
        return redirect(url_for('caixa_sessao_detalhes', sessao_id=sessao.id))

@app.route('/caixa/sessao/<int:sessao_id>/importar/confirmar', methods=['POST'])
@login_required
def confirmar_importacao_movimentos_caixa(sessao_id):
    user_id = session['user_id']
    sessao = CaixaSessao.query.filter_by(id=sessao_id, user_id=user_id).first_or_404()
    data = session.get('csv_import_caixa')
    if not data or data.get('sessao_id') != sessao.id:
        flash('Nenhum CSV em pré-visualização para esta sessão.', 'warning')
        return redirect(url_for('caixa_sessao_detalhes', sessao_id=sessao.id))

    m = data.get('default_map', {})
    mapping = {
        'tipo': request.form.get('map_tipo') or m.get('tipo'),
        'origem': request.form.get('map_origem') or m.get('origem'),
        'valor': request.form.get('map_valor') or m.get('valor'),
        'descricao': request.form.get('map_descricao') or m.get('descricao'),
        'forma_pagamento': request.form.get('map_forma_pagamento') or m.get('forma_pagamento'),
        'referencia_id': request.form.get('map_referencia_id') or m.get('referencia_id'),
        'data': request.form.get('map_data') or m.get('data'),
    }

    imported = 0
    rejected = []
    for row in data.get('rows', []):
        tipo_val = (row.get(mapping['tipo']) if mapping['tipo'] else '').strip().lower()
        origem_val = (row.get(mapping['origem']) if mapping['origem'] else '').strip().lower()
        valor_raw = (row.get(mapping['valor']) if mapping['valor'] else '').strip().replace(',', '.')
        descricao_val = (row.get(mapping['descricao']) if mapping['descricao'] else '').strip()
        forma_val = (row.get(mapping['forma_pagamento']) if mapping['forma_pagamento'] else '').strip()
        ref_val = (row.get(mapping['referencia_id']) if mapping['referencia_id'] else '').strip()

        try:
            valor = float(valor_raw) if valor_raw else None
        except Exception:
            valor = None
        if tipo_val not in ['entrada', 'saida'] or not origem_val or not valor or valor <= 0:
            rejected.append({'row': row, 'reason': 'Campos obrigatórios inválidos'})
            continue

        descricao_final = descricao_val or ''
        if '[IMPORTADO]' not in descricao_final:
            descricao_final = (descricao_final + ' ').strip() + '[IMPORTADO]'

        movimento = MovimentoCaixa(
            tipo=tipo_val,
            origem=origem_val,
            valor=valor,
            descricao=descricao_final,
            forma_pagamento=(forma_val or None),
            referencia_id=int(ref_val) if ref_val and ref_val.isdigit() else None,
            sessao_id=sessao.id,
            user_id=user_id
        )
        db.session.add(movimento)
        imported += 1
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Falha ao confirmar importação: {e}', 'danger')
        return redirect(url_for('caixa_sessao_detalhes', sessao_id=sessao.id))

    session.pop('csv_import_caixa', None)
    session['csv_import_rejeicoes'] = {'count': len(rejected)}
    flash(f'Importação confirmada: {imported} movimentos adicionados, {len(rejected)} rejeições.', 'success')
    return redirect(url_for('caixa_sessao_detalhes', sessao_id=sessao.id))

@app.route('/caixa/sessao/<int:sessao_id>/exportar/<string:formato>')
@login_required
def exportar_caixa_sessao(sessao_id, formato):
    user_id = session['user_id']
    sessao = CaixaSessao.query.filter_by(id=sessao_id, user_id=user_id).first_or_404()

    de_str = request.args.get('de')
    ate_str = request.args.get('ate')
    query = MovimentoCaixa.query.filter_by(sessao_id=sessao.id)
    if de_str:
        try:
            de = datetime.fromisoformat(de_str)
            query = query.filter(MovimentoCaixa.created_at >= de)
        except Exception:
            pass
    if ate_str:
        try:
            ate = datetime.fromisoformat(ate_str)
            query = query.filter(MovimentoCaixa.created_at <= ate)
        except Exception:
            pass
    movimentos = query.order_by(MovimentoCaixa.created_at.asc()).all()

    entradas = sum((m.valor or 0) for m in movimentos if m.tipo == 'entrada')
    saidas = sum((m.valor or 0) for m in movimentos if m.tipo == 'saida')
    saldo_parcial = (sessao.saldo_inicial or 0) + entradas - saidas

    if formato == 'csv':
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Sessão', sessao.id])
        writer.writerow(['Início', sessao.data_abertura.isoformat() if sessao.data_abertura else '-'])
        writer.writerow(['Fim', sessao.data_fechamento.isoformat() if sessao.data_fechamento else '-'])
        writer.writerow([])
        writer.writerow(['Data', 'Tipo', 'Origem', 'Forma pagamento', 'Origem dados', 'Valor', 'Descrição'])
        for m in movimentos:
            origem_dados = 'Importado' if (m.descricao and '[IMPORTADO]' in m.descricao) else 'Manual'
            writer.writerow([
                m.created_at.isoformat() if m.created_at else '',
                m.tipo or '',
                m.origem or '',
                m.forma_pagamento or '',
                origem_dados,
                f'{(m.valor or 0):.2f}',
                m.descricao or ''
            ])
        writer.writerow([])
        writer.writerow(['Entradas', f'{entradas:.2f}'])
        writer.writerow(['Saídas', f'{saidas:.2f}'])
        writer.writerow(['Saldo parcial', f'{saldo_parcial:.2f}'])
        resp = make_response(output.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = f'attachment; filename="caixa_sessao_{sessao.id}.csv"'
        return resp

    elif formato == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'Movimentos'
        ws.append(['Sessão', sessao.id])
        ws.append(['Início', sessao.data_abertura.strftime('%Y-%m-%d %H:%M') if sessao.data_abertura else '-'])
        ws.append(['Fim', sessao.data_fechamento.strftime('%Y-%m-%d %H:%M') if sessao.data_fechamento else '-'])
        ws.append([])
        header = ['Data', 'Tipo', 'Origem', 'Forma pagamento', 'Origem dados', 'Valor', 'Descrição']
        ws.append(header)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        for m in movimentos:
            origem_dados = 'Importado' if (m.descricao and '[IMPORTADO]' in m.descricao) else 'Manual'
            ws.append([
                m.created_at.strftime('%Y-%m-%d %H:%M') if m.created_at else '',
                m.tipo or '',
                m.origem or '',
                m.forma_pagamento or '',
                origem_dados,
                float(f'{(m.valor or 0):.2f}'),
                m.descricao or ''
            ])
        ws.append([])
        ws.append(['Entradas', entradas])
        ws.append(['Saídas', saidas])
        ws.append(['Saldo parcial', saldo_parcial])
        import io
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = make_response(bio.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="caixa_sessao_{sessao.id}.xlsx"'
        return resp

    elif formato == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(f'Sessão do Caixa #{sessao.id}', styles['Title']))
        elements.append(Paragraph(f'Início: {sessao.data_abertura.strftime("%d/%m/%Y %H:%M") if sessao.data_abertura else "-"}', styles['Normal']))
        elements.append(Paragraph(f'Fim: {sessao.data_fechamento.strftime("%d/%m/%Y %H:%M") if sessao.data_fechamento else "-"}', styles['Normal']))
        elements.append(Spacer(1, 12))
        data_tbl = [['Data', 'Tipo', 'Origem', 'Forma pgto.', 'Origem dados', 'Valor', 'Descrição']]
        for m in movimentos:
            origem_dados = 'Importado' if (m.descricao and '[IMPORTADO]' in m.descricao) else 'Manual'
            data_tbl.append([
                m.created_at.strftime('%d/%m/%Y %H:%M') if m.created_at else '',
                m.tipo or '',
                m.origem or '',
                m.forma_pagamento or '',
                origem_dados,
                f'R$ {(m.valor or 0):.2f}',
                (m.descricao or '')[:80]
            ])
        table = Table(data_tbl, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.lightyellow]),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f'Entradas: R$ {entradas:.2f}', styles['Normal']))
        elements.append(Paragraph(f'Saídas: R$ {saidas:.2f}', styles['Normal']))
        elements.append(Paragraph(f'Saldo parcial: R$ {saldo_parcial:.2f}', styles['Normal']))
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        resp = make_response(pdf)
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename="caixa_sessao_{sessao.id}.pdf"'
        return resp

    else:
        flash('Formato de exportação inválido.', 'warning')
        return redirect(url_for('caixa_sessao_detalhes', sessao_id=sessao.id))

@app.route('/api/produtos', methods=['GET', 'POST'])
@csrf.exempt
@jwt_required()
def api_produtos():
    """API completa para produtos"""
    user_id = int(get_jwt_identity())
    
    if request.method == 'GET':
        # Listar produtos com paginação
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 20, type=int), 100)
        search = request.args.get('search', '', type=str)
        categoria = request.args.get('categoria', '', type=str)
        
        query = Produto.query.filter_by(user_id=user_id)
        
        # Filtros
        if search:
            query = query.filter(Produto.nome.contains(search))
        if categoria:
            query = query.filter_by(categoria=categoria)
        
        # Paginação
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        produtos = pagination.items
        
        return jsonify({
            'produtos': [{
                'id': p.id,
                'nome': p.nome,
                'descricao': p.descricao,
                'preco': float(p.preco),
                'estoque_atual': p.estoque_atual,
                'estoque_minimo': p.estoque_minimo,
                'categoria': p.categoria,
                'codigo_barras': p.codigo_barras,
                'created_at': p.created_at.isoformat() if p.created_at else None
            } for p in produtos],
            'pagination': {
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        })
    
    elif request.method == 'POST':
        # Criar novo produto
        data = request.get_json()
        
        if not data or not data.get('nome'):
            return jsonify({'error': 'Nome do produto é obrigatório'}), 400
        
        produto = Produto(
            nome=sanitize_input(data['nome']),
            descricao=sanitize_input(data.get('descricao', '')),
            preco=float(data.get('preco', 0)),
            estoque_atual=int(data.get('estoque_atual', 0)),
            estoque_minimo=int(data.get('estoque_minimo', 0)),
            categoria=sanitize_input(data.get('categoria', '')),
            codigo_barras=sanitize_input(data.get('codigo_barras', '')),
            user_id=user_id
        )
        
        try:
            db.session.add(produto)
            db.session.commit()
            
            return jsonify({
                'id': produto.id,
                'nome': produto.nome,
                'descricao': produto.descricao,
                'preco': float(produto.preco),
                'estoque_atual': produto.estoque_atual,
                'estoque_minimo': produto.estoque_minimo,
                'categoria': produto.categoria,
                'codigo_barras': produto.codigo_barras,
                'created_at': produto.created_at.isoformat()
            }), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'Erro ao criar produto'}), 500

@app.route('/api/produtos/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@csrf.exempt
@jwt_required()
def api_produto(id):
    """API para produto específico"""
    user_id = int(get_jwt_identity())
    produto = Produto.query.filter_by(id=id, user_id=user_id).first()
    
    if not produto:
        return jsonify({'error': 'Produto não encontrado'}), 404
    
    if request.method == 'GET':
        return jsonify({
            'id': produto.id,
            'nome': produto.nome,
            'descricao': produto.descricao,
            'preco': float(produto.preco),
            'estoque_atual': produto.estoque_atual,
            'estoque_minimo': produto.estoque_minimo,
            'categoria': produto.categoria,
            'codigo_barras': produto.codigo_barras,
            'created_at': produto.created_at.isoformat() if produto.created_at else None
        })
    
    elif request.method == 'PUT':
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Dados são obrigatórios'}), 400
        
        # Atualizar campos fornecidos
        if 'nome' in data:
            produto.nome = sanitize_input(data['nome'])
        if 'descricao' in data:
            produto.descricao = sanitize_input(data['descricao'])
        if 'preco' in data:
            produto.preco = float(data['preco'])
        if 'estoque_atual' in data:
            produto.estoque_atual = int(data['estoque_atual'])
        if 'estoque_minimo' in data:
            produto.estoque_minimo = int(data['estoque_minimo'])
        if 'categoria' in data:
            produto.categoria = sanitize_input(data['categoria'])
        if 'codigo_barras' in data:
            produto.codigo_barras = sanitize_input(data['codigo_barras'])
        
        try:
            db.session.commit()
            
            return jsonify({
                'id': produto.id,
                'nome': produto.nome,
                'descricao': produto.descricao,
                'preco': float(produto.preco),
                'estoque_atual': produto.estoque_atual,
                'estoque_minimo': produto.estoque_minimo,
                'categoria': produto.categoria,
                'codigo_barras': produto.codigo_barras,
                'created_at': produto.created_at.isoformat() if produto.created_at else None
            }), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'Erro ao atualizar produto'}), 500
    
    elif request.method == 'DELETE':
        try:
            db.session.delete(produto)
            db.session.commit()
            return jsonify({'message': 'Produto removido com sucesso'}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'Erro ao remover produto'}), 500

# API para clientes
@app.route('/api/clientes', methods=['GET', 'POST'])
@csrf.exempt
@jwt_required()
def api_clientes():
    """API completa para clientes"""
    user_id = int(get_jwt_identity())
    
    if request.method == 'GET':
        # Listar clientes com paginação
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 20, type=int), 100)
        search = request.args.get('search', '', type=str)
        
        query = Cliente.query.filter_by(user_id=user_id)
        
        # Filtros
        if search:
            query = query.filter(Cliente.nome.contains(search))
        
        # Paginação
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        clientes = pagination.items
        
        return jsonify({
            'clientes': [{
                'id': c.id,
                'nome': c.nome,
                'email': c.email,
                'telefone': c.telefone,
                'endereco': c.endereco,
                'cidade': c.cidade,
                'estado': c.estado,
                'cep': c.cep,
                'created_at': c.created_at.isoformat() if c.created_at else None
            } for c in clientes],
            'pagination': {
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        })
    
    elif request.method == 'POST':
        # Criar novo cliente
        data = request.get_json()
        
        if not data or not data.get('nome'):
            return jsonify({'error': 'Nome do cliente é obrigatório'}), 400
        
        cliente = Cliente(
            nome=sanitize_input(data['nome']),
            email=sanitize_input(data.get('email', '')),
            telefone=sanitize_input(data.get('telefone', '')),
            endereco=sanitize_input(data.get('endereco', '')),
            user_id=user_id
        )
        
        try:
            db.session.add(cliente)
            db.session.commit()
            
            return jsonify({
                'id': cliente.id,
                'nome': cliente.nome,
                'email': cliente.email,
                'telefone': cliente.telefone,
                'endereco': cliente.endereco,
                'cidade': cliente.cidade,
                'estado': cliente.estado,
                'cep': cliente.cep,
                'created_at': cliente.created_at.isoformat()
            }), 201
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'Erro ao criar cliente'}), 500

@app.route('/api/clientes/<int:id>', methods=['GET', 'PUT', 'DELETE'])
@csrf.exempt
@jwt_required()
def api_cliente(id):
    """API para cliente específico"""
    user_id = int(get_jwt_identity())
    cliente = Cliente.query.filter_by(id=id, user_id=user_id).first()
    
    if not cliente:
        return jsonify({'error': 'Cliente não encontrado'}), 404
    
    if request.method == 'GET':
        return jsonify({
            'id': cliente.id,
            'nome': cliente.nome,
            'email': cliente.email,
            'telefone': cliente.telefone,
            'endereco': cliente.endereco,
            'cidade': cliente.cidade,
            'estado': cliente.estado,
            'cep': cliente.cep,
            'created_at': cliente.created_at.isoformat() if cliente.created_at else None
        })
    
    elif request.method == 'PUT':
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Dados são obrigatórios'}), 400
        
        # Atualizar campos fornecidos
        if 'nome' in data:
            cliente.nome = sanitize_input(data['nome'])
        if 'email' in data:
            cliente.email = sanitize_input(data['email'])
        if 'telefone' in data:
            cliente.telefone = sanitize_input(data['telefone'])
        if 'endereco' in data:
            cliente.endereco = sanitize_input(data['endereco'])
        if 'cidade' in data:
            cliente.cidade = sanitize_input(data['cidade'])
        if 'estado' in data:
            cliente.estado = sanitize_input(data['estado'])
        if 'cep' in data:
            cliente.cep = sanitize_input(data['cep'])
        
        try:
            db.session.commit()
            
            return jsonify({
                'id': cliente.id,
                'nome': cliente.nome,
                'email': cliente.email,
                'telefone': cliente.telefone,
                'endereco': cliente.endereco,
                'cidade': cliente.cidade,
                'estado': cliente.estado,
                'cep': cliente.cep,
                'created_at': cliente.created_at.isoformat() if cliente.created_at else None
            }), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'Erro ao atualizar cliente'}), 500
    
    elif request.method == 'DELETE':
        try:
            db.session.delete(cliente)
            db.session.commit()
            return jsonify({'message': 'Cliente removido com sucesso'}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'Erro ao remover cliente'}), 500

# API para buscar fornecedores
@app.route('/api/fornecedores')
@csrf.exempt
@login_required
def api_fornecedores():
    user_id = session['user_id']
    fornecedores = Fornecedor.query.filter_by(user_id=user_id, status='ativo').all()
    
    return jsonify([{
        'id': f.id,
        'nome': f.nome,
        'razao_social': f.razao_social,
        'cnpj': f.cnpj
    } for f in fornecedores])

# API para categorias
@app.route('/api/categorias', methods=['GET', 'POST'])
@csrf.exempt
@login_required
def api_categorias():
    if request.method == 'GET':
        # Buscar todas as categorias existentes no sistema (produtos e produtos auxiliares)
        user_id = session['user_id']
        
        # Categorias de produtos
        categorias_produtos = db.session.query(Produto.categoria).filter(
            Produto.user_id == user_id,
            Produto.categoria.isnot(None),
            Produto.categoria != ''
        ).distinct().all()
        
        # Categorias de produtos auxiliares
        categorias_auxiliares = db.session.query(ProdutoAuxiliar.categoria).filter(
            ProdutoAuxiliar.user_id == user_id,
            ProdutoAuxiliar.categoria.isnot(None),
            ProdutoAuxiliar.categoria != ''
        ).distinct().all()
        
        # Combinar e remover duplicatas
        todas_categorias = set()
        todas_categorias.update([cat[0] for cat in categorias_produtos])
        todas_categorias.update([cat[0] for cat in categorias_auxiliares])
        
        lista_categorias = sorted(list(todas_categorias))
        
        return jsonify(lista_categorias)
    
    elif request.method == 'POST':
        # Criar nova categoria com verificação de duplicação
        data = request.get_json()
        nova_categoria = data.get('categoria', '').strip()
        
        if not nova_categoria:
            return jsonify({'success': False, 'error': 'Nome da categoria é obrigatório'}), 400
        
        user_id = session['user_id']
        
        # Verificar se categoria já existe (case-insensitive)
        categoria_existe = db.session.query(Produto.categoria).filter(
            Produto.user_id == user_id,
            db.func.lower(Produto.categoria) == nova_categoria.lower()
        ).first()
        
        if not categoria_existe:
            categoria_existe = db.session.query(ProdutoAuxiliar.categoria).filter(
                ProdutoAuxiliar.user_id == user_id,
                db.func.lower(ProdutoAuxiliar.categoria) == nova_categoria.lower()
            ).first()
        
        if categoria_existe:
            return jsonify({
                'success': False, 
                'error': f'A categoria "{nova_categoria}" já existe no sistema'
            }), 409
        
        # Categoria não existe, pode ser criada
        return jsonify({
            'success': True, 
            'categoria': nova_categoria,
            'message': f'Categoria "{nova_categoria}" pode ser criada'
        })

# Rotas de Configurações
@app.route('/configuracoes')
@login_required
def configuracoes():
    user_id = session['user_id']
    settings = get_user_settings(user_id)
    return render_template('configuracoes/index.html', settings=settings)

@app.route('/configuracoes/atualizar', methods=['POST'])
@login_required
def atualizar_configuracoes():
    user_id = session['user_id']
    
    # Processar dados do formulário
    dark_mode = request.form.get('dark_mode') == 'on'
    notifications = request.form.get('notifications') == 'on'
    auto_logout = int(request.form.get('auto_logout', 30))
    language = request.form.get('language', 'pt')
    timezone = request.form.get('timezone', 'America/Sao_Paulo')
    dashboard_refresh = int(request.form.get('dashboard_refresh', 30))
    
    # Debug: imprimir valores recebidos
    print(f"Dark mode received: {request.form.get('dark_mode')}")
    print(f"Dark mode processed: {dark_mode}")
    print(f"User ID: {user_id}")
    
    # Atualizar configurações
    update_user_settings(user_id,
                        dark_mode=dark_mode,
                        notifications=notifications,
                        auto_logout=auto_logout,
                        language=language,
                        timezone=timezone,
                        dashboard_refresh=dashboard_refresh)
    
    # Atualizar sessão
    session['dark_mode'] = dark_mode
    
    flash('Configurações atualizadas com sucesso!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/api/toggle-dark-mode', methods=['POST'])
@csrf.exempt
@login_required
def toggle_dark_mode():
    user_id = session['user_id']
    settings = get_user_settings(user_id)
    
    # Alternar dark mode
    new_dark_mode = not settings.dark_mode
    update_user_settings(user_id, dark_mode=new_dark_mode)
    
    # Atualizar sessão
    session['dark_mode'] = new_dark_mode
    
    return jsonify({
        'success': True,
        'dark_mode': new_dark_mode,
        'message': 'Modo escuro ' + ('ativado' if new_dark_mode else 'desativado')
    })

# ==================== RELATÓRIOS E ANALYTICS ====================

# Página principal de relatórios
@app.route('/relatorios')
@login_required
def relatorios():
    user_id = session['user_id']
    
    # KPIs rápidos para o dashboard de relatórios
    total_produtos = Produto.query.filter_by(user_id=user_id).count()
    total_clientes = Cliente.query.filter_by(user_id=user_id).count()
    
    # Vendas do mês atual
    inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    vendas_mes = Venda.query.filter(
        Venda.user_id == user_id,
        Venda.data_venda >= inicio_mes,
        Venda.status == 'finalizada'
    ).all()
    
    receita_mes = sum(venda.valor_total for venda in vendas_mes)
    
    # Produtos com estoque baixo
    produtos_estoque_baixo = Produto.query.filter(
        Produto.user_id == user_id,
        Produto.estoque_atual <= Produto.estoque_minimo
    ).count()
    
    return render_template('relatorios/index.html',
                         total_produtos=total_produtos,
                         total_clientes=total_clientes,
                         receita_mes=receita_mes,
                         produtos_estoque_baixo=produtos_estoque_baixo)

# ==================== RELATÓRIOS FINANCEIROS ====================

# P&L (Profit & Loss)
@app.route('/relatorios/pl')
@login_required
def relatorio_pl():
    user_id = session['user_id']
    
    # Período (padrão: mês atual)
    inicio = request.args.get('inicio', datetime.now().replace(day=1).strftime('%Y-%m-%d'))
    fim = request.args.get('fim', datetime.now().strftime('%Y-%m-%d'))
    
    # Receitas
    vendas = Venda.query.filter(
        Venda.user_id == user_id,
        Venda.data_venda.between(inicio, fim),
        Venda.status == 'finalizada'
    ).all()
    
    receita_total = sum(v.valor_total for v in vendas)
    
    # Custos
    compras = Compra.query.filter(
        Compra.user_id == user_id,
        Compra.data_compra.between(inicio, fim)
    ).all()
    
    custo_total = sum(c.valor_total for c in compras)
    
    # Margem de lucro
    margem_lucro = receita_total - custo_total
    margem_percentual = (margem_lucro / receita_total * 100) if receita_total > 0 else 0
    
    return render_template('relatorios/pl.html', 
                         receita_total=receita_total,
                         custo_total=custo_total,
                         margem_lucro=margem_lucro,
                         margem_percentual=margem_percentual,
                         periodo={'inicio': inicio, 'fim': fim})

# Fluxo de Caixa
@app.route('/relatorios/fluxo-caixa')
@login_required
def relatorio_fluxo_caixa():
    user_id = session['user_id']
    
    # Últimos 12 meses
    fluxo_mensal = []
    for i in range(12):
        mes = datetime.now() - timedelta(days=30*i)
        inicio_mes = mes.replace(day=1)
        fim_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Entradas (vendas)
        vendas_mes = Venda.query.filter(
            Venda.user_id == user_id,
            Venda.data_venda.between(inicio_mes, fim_mes),
            Venda.status == 'finalizada'
        ).all()
        
        entradas = sum(v.valor_total for v in vendas_mes)
        
        # Saídas (compras)
        compras_mes = Compra.query.filter(
            Compra.user_id == user_id,
            Compra.data_compra.between(inicio_mes, fim_mes)
        ).all()
        
        saidas = sum(c.valor_total for c in compras_mes)
        
        fluxo_mensal.append({
            'mes': inicio_mes.strftime('%Y-%m'),
            'entradas': entradas,
            'saidas': saidas,
            'saldo': entradas - saidas
        })
    
    return render_template('relatorios/fluxo_caixa.html', fluxo_mensal=fluxo_mensal)

# ==================== ANÁLISE DE VENDAS ====================

# Top Produtos
@app.route('/relatorios/top-produtos')
@login_required
def relatorio_top_produtos():
    user_id = session['user_id']
    
    # Query para produtos mais vendidos
    top_produtos = db.session.query(
        Produto.nome,
        db.func.sum(ItemVenda.quantidade).label('total_vendido'),
        db.func.sum(ItemVenda.quantidade * ItemVenda.preco_unitario).label('receita_total')
    ).join(ItemVenda).join(Venda).filter(
        Venda.user_id == user_id,
        Venda.status == 'finalizada'
    ).group_by(Produto.id, Produto.nome).order_by(
        db.func.sum(ItemVenda.quantidade).desc()
    ).limit(10).all()
    
    return render_template('relatorios/top_produtos.html', top_produtos=top_produtos)

# Análise de Sazonalidade
@app.route('/relatorios/sazonalidade')
@login_required
def relatorio_sazonalidade():
    user_id = session['user_id']
    
    # Vendas por mês (últimos 12 meses)
    vendas_por_mes = []
    for i in range(12):
        mes = datetime.now() - timedelta(days=30*i)
        inicio_mes = mes.replace(day=1)
        fim_mes = (inicio_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        vendas_mes = Venda.query.filter(
            Venda.user_id == user_id,
            Venda.data_venda.between(inicio_mes, fim_mes),
            Venda.status == 'finalizada'
        ).all()
        
        vendas_por_mes.append({
            'mes': inicio_mes.strftime('%Y-%m'),
            'total': sum(v.valor_total for v in vendas_mes),
            'quantidade': len(vendas_mes)
        })
    
    return render_template('relatorios/sazonalidade.html', vendas_por_mes=vendas_por_mes)

# ==================== RELATÓRIOS DE ESTOQUE ====================

# Rotatividade de Produtos
@app.route('/relatorios/rotatividade-estoque')
@login_required
def relatorio_rotatividade_estoque():
    user_id = session['user_id']
    
    produtos_rotatividade = []
    produtos = Produto.query.filter_by(user_id=user_id).all()
    
    for produto in produtos:
        # Vendas do produto nos últimos 30 dias
        inicio = datetime.now() - timedelta(days=30)
        vendas_produto = db.session.query(
            db.func.sum(ItemVenda.quantidade)
        ).join(Venda).filter(
            ItemVenda.produto_id == produto.id,
            Venda.user_id == user_id,
            Venda.data_venda >= inicio,
            Venda.status == 'finalizada'
        ).scalar() or 0
        
        # Rotatividade = vendas / estoque médio
        estoque_medio = produto.estoque_atual
        rotatividade = (vendas_produto / estoque_medio) if estoque_medio > 0 else 0
        
        produtos_rotatividade.append({
            'produto': produto,
            'vendas_30_dias': vendas_produto,
            'estoque_atual': produto.estoque_atual,
            'rotatividade': rotatividade
        })
    
    # Ordenar por rotatividade
    produtos_rotatividade.sort(key=lambda x: x['rotatividade'], reverse=True)
    
    return render_template('relatorios/rotatividade_estoque.html', 
                         produtos_rotatividade=produtos_rotatividade)

# Produtos Parados
@app.route('/relatorios/produtos-parados')
@login_required
def relatorio_produtos_parados():
    user_id = session['user_id']
    
    # Produtos sem vendas nos últimos 90 dias
    inicio = datetime.now() - timedelta(days=90)
    
    produtos_parados = []
    produtos = Produto.query.filter_by(user_id=user_id).all()
    
    for produto in produtos:
        # Verificar se teve vendas nos últimos 90 dias
        vendas_recentes = db.session.query(ItemVenda).join(Venda).filter(
            ItemVenda.produto_id == produto.id,
            Venda.user_id == user_id,
            Venda.data_venda >= inicio,
            Venda.status == 'finalizada'
        ).first()
        
        if not vendas_recentes and produto.estoque_atual > 0:
            # Última venda
            ultima_venda = db.session.query(Venda.data_venda).join(ItemVenda).filter(
                ItemVenda.produto_id == produto.id,
                Venda.user_id == user_id,
                Venda.status == 'finalizada'
            ).order_by(Venda.data_venda.desc()).first()
            
            produtos_parados.append({
                'produto': produto,
                'ultima_venda': ultima_venda[0] if ultima_venda else None,
                'dias_sem_venda': (datetime.now() - ultima_venda[0]).days if ultima_venda else 999
            })
    
    # Ordenar por dias sem venda
    produtos_parados.sort(key=lambda x: x['dias_sem_venda'], reverse=True)
    
    return render_template('relatorios/produtos_parados.html', 
                         produtos_parados=produtos_parados)

# ==================== DASHBOARD EXECUTIVO ====================

# API para KPIs em tempo real
@app.route('/api/dashboard-kpis')
@csrf.exempt
@login_required
def api_dashboard_kpis():
    user_id = session['user_id']
    
    # Período atual vs anterior
    hoje = datetime.now().date()
    inicio_mes_atual = hoje.replace(day=1)
    inicio_mes_anterior = (inicio_mes_atual - timedelta(days=1)).replace(day=1)
    fim_mes_anterior = inicio_mes_atual - timedelta(days=1)
    
    # Vendas do mês atual
    vendas_atual = Venda.query.filter(
        Venda.user_id == user_id,
        Venda.data_venda >= inicio_mes_atual,
        Venda.status == 'finalizada'
    ).all()
    
    # Vendas do mês anterior
    vendas_anterior = Venda.query.filter(
        Venda.user_id == user_id,
        Venda.data_venda.between(inicio_mes_anterior, fim_mes_anterior),
        Venda.status == 'finalizada'
    ).all()
    
    receita_atual = sum(v.valor_total for v in vendas_atual)
    receita_anterior = sum(v.valor_total for v in vendas_anterior)
    
    # Crescimento percentual
    crescimento = ((receita_atual - receita_anterior) / receita_anterior * 100) if receita_anterior > 0 else 0
    
    # Outros KPIs
    total_clientes = Cliente.query.filter_by(user_id=user_id).count()
    produtos_estoque_baixo = Produto.query.filter(
        Produto.user_id == user_id,
        Produto.estoque_atual <= Produto.estoque_minimo
    ).count()
    
    return jsonify({
        'receita_atual': receita_atual,
        'receita_anterior': receita_anterior,
        'crescimento': crescimento,
        'total_clientes': total_clientes,
        'produtos_estoque_baixo': produtos_estoque_baixo,
        'ticket_medio': receita_atual / len(vendas_atual) if vendas_atual else 0
    })

# ==================== SISTEMA DE EXPORTAÇÃO ====================

# Exportação para PDF
@app.route('/relatorios/exportar/<tipo>/<formato>')
@login_required
def exportar_relatorio(tipo, formato):
    user_id = session['user_id']
    
    if formato == 'pdf':
        return exportar_pdf(tipo, user_id)
    elif formato == 'excel':
        return exportar_excel(tipo, user_id)
    elif formato == 'csv':
        return exportar_csv(tipo, user_id)

def exportar_pdf(tipo, user_id):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title = Paragraph(f"Relatório de {tipo.title()}", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Dados baseados no tipo
    if tipo == 'vendas':
        vendas = Venda.query.filter_by(user_id=user_id).all()
        data = [['Data', 'Cliente', 'Valor', 'Status']]
        for venda in vendas:
            data.append([
                venda.data_venda.strftime('%d/%m/%Y'),
                venda.cliente.nome if venda.cliente else 'N/A',
                f"R$ {venda.valor_total:.2f}",
                venda.status
            ])
    elif tipo == 'produtos':
        produtos = Produto.query.filter_by(user_id=user_id).all()
        data = [['Nome', 'Preço', 'Estoque', 'Categoria']]
        for produto in produtos:
            data.append([
                produto.nome,
                f"R$ {produto.preco:.2f}",
                str(produto.estoque_atual),
                produto.categoria or 'N/A'
            ])
    
    # Criar tabela
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_{tipo}.pdf'
    
    return response

def exportar_excel(tipo, user_id):
    buffer = BytesIO()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Relatório {tipo.title()}"
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados baseados no tipo
    if tipo == 'vendas':
        vendas = Venda.query.filter_by(user_id=user_id).all()
        headers = ['Data', 'Cliente', 'Valor', 'Status']
        ws.append(headers)
        
        for venda in vendas:
            ws.append([
                venda.data_venda.strftime('%d/%m/%Y'),
                venda.cliente.nome if venda.cliente else 'N/A',
                venda.valor_total,
                venda.status
            ])
    elif tipo == 'produtos':
        produtos = Produto.query.filter_by(user_id=user_id).all()
        headers = ['Nome', 'Preço', 'Estoque', 'Categoria']
        ws.append(headers)
        
        for produto in produtos:
            ws.append([
                produto.nome,
                produto.preco,
                produto.estoque_atual,
                produto.categoria or 'N/A'
            ])
    
    # Aplicar estilo ao cabeçalho
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_{tipo}.xlsx'
    
    return response

def exportar_csv(tipo, user_id):
    buffer = BytesIO()
    
    # Dados baseados no tipo
    if tipo == 'vendas':
        vendas = Venda.query.filter_by(user_id=user_id).all()
        data = []
        for venda in vendas:
            data.append({
                'Data': venda.data_venda.strftime('%d/%m/%Y'),
                'Cliente': venda.cliente.nome if venda.cliente else 'N/A',
                'Valor': venda.valor_total,
                'Status': venda.status
            })
    elif tipo == 'produtos':
        produtos = Produto.query.filter_by(user_id=user_id).all()
        data = []
        for produto in produtos:
            data.append({
                'Nome': produto.nome,
                'Preço': produto.preco,
                'Estoque': produto.estoque_atual,
                'Categoria': produto.categoria or 'N/A'
            })
    
    # Converter para CSV usando implementação nativa do Python
    import csv
    import io
    
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    csv_string = output.getvalue()
    output.close()
    
    response = make_response(csv_string)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_{tipo}.csv'
    
    return response

# Rotas de Cupons
@app.route('/cupons')
@login_required
def cupons():
    user_id = session['user_id']
    cupons = Cupom.query.filter_by(user_id=user_id).order_by(Cupom.created_at.desc()).all()
    return render_template('cupons/list.html', cupons=cupons)

@app.route('/cupons/novo', methods=['GET', 'POST'])
@login_required
def novo_cupom():
    if request.method == 'POST':
        try:
            # Validar campos obrigatórios
            if not request.form.get('codigo'):
                flash('Código do cupom é obrigatório!', 'error')
                return render_template('cupons/form.html')
            
            if not request.form.get('valor_desconto'):
                flash('Valor do desconto é obrigatório!', 'error')
                return render_template('cupons/form.html')
            
            # Verificar se código já existe
            codigo = request.form['codigo'].upper().strip()
            cupom_existente = Cupom.query.filter_by(codigo=codigo).first()
            if cupom_existente:
                flash('Código de cupom já existe!', 'error')
                return render_template('cupons/form.html')
            
            # Processar dados
            valor_desconto = float(request.form['valor_desconto'].replace(',', '.'))
            valor_minimo_compra = float(request.form.get('valor_minimo_compra', 0).replace(',', '.'))
            valor_maximo_desconto = None
            if request.form.get('valor_maximo_desconto'):
                valor_maximo_desconto = float(request.form['valor_maximo_desconto'].replace(',', '.'))
            
            limite_uso = None
            if request.form.get('limite_uso'):
                limite_uso = int(request.form['limite_uso'])
            
            # Processar período de validade (opcional)
            data_inicio = None
            data_fim = None
            tem_periodo_validade = request.form.get('tem_periodo_validade') == 'on'
            
            if tem_periodo_validade:
                if not request.form.get('data_inicio') or not request.form.get('data_fim'):
                    flash('Quando período de validade está habilitado, as datas são obrigatórias!', 'error')
                    return render_template('cupons/form.html')
                
                from datetime import datetime
                data_inicio = datetime.strptime(request.form['data_inicio'], '%Y-%m-%d')
                data_fim = datetime.strptime(request.form['data_fim'], '%Y-%m-%d')
                
                # Validar datas
                if data_fim <= data_inicio:
                    flash('Data de fim deve ser posterior à data de início!', 'error')
                    return render_template('cupons/form.html')
            
            # Criar cupom
            cupom = Cupom(
                codigo=codigo,
                descricao=request.form.get('descricao', '').strip(),
                tipo_desconto=request.form['tipo_desconto'],
                valor_desconto=valor_desconto,
                valor_minimo_compra=valor_minimo_compra,
                valor_maximo_desconto=valor_maximo_desconto,
                limite_uso=limite_uso,
                data_inicio=data_inicio,
                data_fim=data_fim,
                ativo=request.form.get('ativo') == 'on',
                user_id=session['user_id']
            )
            
            db.session.add(cupom)
            db.session.commit()
            
            flash('Cupom criado com sucesso!', 'success')
            return redirect(url_for('cupons'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar cupom: {str(e)}', 'error')
            return render_template('cupons/form.html')
    
    return render_template('cupons/form.html')

@app.route('/cupons/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cupom(id):
    cupom = Cupom.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    if request.method == 'POST':
        try:
            # Verificar se código já existe (exceto o próprio cupom)
            codigo = request.form['codigo'].upper().strip()
            cupom_existente = Cupom.query.filter(Cupom.codigo == codigo, Cupom.id != id).first()
            if cupom_existente:
                flash('Código de cupom já existe!', 'error')
                return render_template('cupons/form.html', cupom=cupom)
            
            # Processar dados
            valor_desconto = float(request.form['valor_desconto'].replace(',', '.'))
            valor_minimo_compra = float(request.form.get('valor_minimo_compra', 0).replace(',', '.'))
            valor_maximo_desconto = None
            if request.form.get('valor_maximo_desconto'):
                valor_maximo_desconto = float(request.form['valor_maximo_desconto'].replace(',', '.'))
            
            limite_uso = None
            if request.form.get('limite_uso'):
                limite_uso = int(request.form['limite_uso'])
            
            # Processar período de validade (opcional)
            data_inicio = None
            data_fim = None
            tem_periodo_validade = request.form.get('tem_periodo_validade') == 'on'
            
            if tem_periodo_validade:
                if not request.form.get('data_inicio') or not request.form.get('data_fim'):
                    flash('Quando período de validade está habilitado, as datas são obrigatórias!', 'error')
                    return render_template('cupons/form.html', cupom=cupom)
                
                from datetime import datetime
                data_inicio = datetime.strptime(request.form['data_inicio'], '%Y-%m-%d')
                data_fim = datetime.strptime(request.form['data_fim'], '%Y-%m-%d')
                
                # Validar datas
                if data_fim <= data_inicio:
                    flash('Data de fim deve ser posterior à data de início!', 'error')
                    return render_template('cupons/form.html', cupom=cupom)
            
            # Atualizar cupom
            cupom.codigo = codigo
            cupom.descricao = request.form.get('descricao', '').strip()
            cupom.tipo_desconto = request.form['tipo_desconto']
            cupom.valor_desconto = valor_desconto
            cupom.valor_minimo_compra = valor_minimo_compra
            cupom.valor_maximo_desconto = valor_maximo_desconto
            cupom.limite_uso = limite_uso
            cupom.data_inicio = data_inicio
            cupom.data_fim = data_fim
            cupom.ativo = request.form.get('ativo') == 'on'
            
            db.session.commit()
            
            flash('Cupom atualizado com sucesso!', 'success')
            return redirect(url_for('cupons'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar cupom: {str(e)}', 'error')
            return render_template('cupons/form.html', cupom=cupom)
    
    return render_template('cupons/form.html', cupom=cupom)

@app.route('/cupons/excluir/<int:id>')
@login_required
def excluir_cupom(id):
    cupom = Cupom.query.filter_by(id=id, user_id=session['user_id']).first_or_404()
    
    # Verificar se cupom foi usado em vendas
    if cupom.usos_realizados > 0:
        flash('Não é possível excluir cupom que já foi utilizado!', 'error')
        return redirect(url_for('cupons'))
    
    db.session.delete(cupom)
    db.session.commit()
    
    flash('Cupom excluído com sucesso!', 'success')
    return redirect(url_for('cupons'))

@app.route('/api/cupom/<string:codigo>')
@login_required
def validar_cupom(codigo):
    """API para validar cupom e retornar informações"""
    from datetime import datetime
    
    cupom = Cupom.query.filter_by(codigo=codigo.upper(), user_id=session['user_id']).first()
    
    if not cupom:
        return jsonify({'valido': False, 'mensagem': 'Cupom não encontrado'})
    
    # Verificar se cupom está ativo
    if not cupom.ativo:
        return jsonify({'valido': False, 'mensagem': 'Cupom inativo'})
    
    # Verificar datas (apenas se cupom tiver período de validade definido)
    if cupom.data_inicio and cupom.data_fim:
        agora = datetime.utcnow()
        if agora < cupom.data_inicio:
            return jsonify({'valido': False, 'mensagem': 'Cupom ainda não é válido'})
        
        if agora > cupom.data_fim:
            return jsonify({'valido': False, 'mensagem': 'Cupom expirado'})
    
    # Verificar limite de uso
    if cupom.limite_uso and cupom.usos_realizados >= cupom.limite_uso:
        return jsonify({'valido': False, 'mensagem': 'Cupom esgotado'})
    
    # Retornar informações do cupom
    return jsonify({
        'valido': True,
        'cupom': {
            'id': cupom.id,
            'codigo': cupom.codigo,
            'descricao': cupom.descricao,
            'tipo_desconto': cupom.tipo_desconto,
            'valor_desconto': cupom.valor_desconto,
            'valor_minimo_compra': cupom.valor_minimo_compra,
            'valor_maximo_desconto': cupom.valor_maximo_desconto
        }
    })

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
